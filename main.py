"""
Carousell Auto Listing Script (v2.0)
Automates creating and editing listings on Carousell Android app
Uses adbutils for device control, opencv for image matching, and EasyOCR for text recognition
"""

import os
import sys
import time
import random
from datetime import datetime
from pathlib import Path
import numpy as np
import cv2
import pandas as pd
from openpyxl import load_workbook
from adbutils import adb
import subprocess
import socket
import shlex

# Import device manager for Smart Hybrid Input
from device_manager import (
    set_screen_dimensions,
    get_device_resolution,
    pixels_to_rect,
    SCREEN_WIDTH,
    SCREEN_HEIGHT,
    input_text_stealth, 
    human_click_with_pressure,
    clear_text_native,
    smart_input,
    swipe_down_large_distance_in_pixel_region
)

# Import EasyOCR if available
try:
    import easyocr
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    print("⚠ EasyOCR not installed. OCR features disabled.")

# Configuration
ASSETS_DIR = r"D:\Carousell_Auto\assets"
EXCEL_PATH = r"D:\Carousell_Auto\data\output_007_replaced_updated.xlsx"
PC_IMAGES_BASE = r"D:\Carousell_Auto\data"
PHONE_TEMP_DIR = "/sdcard/Pictures/Album_Cache_Temp"
APP_PACKAGE = "com.thecarousell.Carousell"

SCREENSHOT_THRESHOLD = 0.8
DEFAULT_TIMEOUT = 10
WAIT_AFTER_CLICK = 1.5
WAIT_AFTER_INPUT = 1.0
APP_RESTART_WAIT = 12

# Coordinate Configuration (Percentage-based)
# Sell Button (bottom center) - STRICT BOUNDS
SELL_BUTTON_RECT = {
    'x_min': 0.475,
    'x_max': 0.522,
    'y_min': 0.924,
    'y_max': 0.945
}

# Manage Button (top-right, only appears when drafts exist)
MANAGE_BUTTON = {
    'x_min': 0.803,
    'x_max': 0.958,
    'y_min': 0.136,
    'y_max': 0.151
}

# Folder Dropdown (top-right area) - CORRECTED BOUNDS
DROPDOWN_RECT = {
    'x_min': 0.814,
    'x_max': 0.956,
    'y_min': 0.139,
    'y_max': 0.150
}

# Target Folder - Dynamic (based on "Photos" text anchor)
# No longer using fixed FOLDER_RECT, will calculate from Photos position

# Draft cleanup - Grid positions (2x3 grid, 6 items)
DRAFT_GRID = {
    'positions': [
        {'x_min': 0.050, 'x_max': 0.431, 'y_min': 0.158, 'y_max': 0.330},  # Left 1
        {'x_min': 0.506, 'x_max': 0.892, 'y_min': 0.159, 'y_max': 0.330},  # Right 1
        {'x_min': 0.047, 'x_max': 0.428, 'y_min': 0.383, 'y_max': 0.550},  # Left 2
        # Add more if needed
    ],
    'avoid_delete': {'y_min': 0.909, 'y_max': 0.955}  # Delete button zone to avoid
}

# Delete button (bottom)
DELETE_BUTTON = {
    'x_min': 0.050,
    'x_max': 0.953,
    'y_min': 0.909,
    'y_max': 0.955
}

# Image Grid (3x5, skip camera)
IMAGE_GRID = {
    'top_left': (0.033, 0.198),
    'bottom_right': (1.012, 0.940),
    'columns': 3,
    'rows': 5,
    'skip_first': True,
    'cell_jitter': 0.03
}

# Next Button
NEXT_BUTTON = {
    'x_min': 0.056,
    'x_max': 0.947,
    'y_min': 0.915,
    'y_max': 0.959
}

# Category Input Box (based on "Category" text position)
CATEGORY_TEXT_CENTER = (0.186, 0.383)  # "Category" text center
CATEGORY_INPUT_BOX = {
    'x_min': 0.083,
    'x_max': 0.919,
    'y_min': 0.409,
    'y_max': 0.491
}

# Initialize device
print("Connecting to Android device...")
d = adb.device()
print(f"Device connected: {d.serial}")

# Get device resolution using ADB (more reliable than window_size)
print("\n📱 Detecting device screen resolution...")
screen_width, screen_height = get_device_resolution()
print(f"Screen resolution: {screen_width}x{screen_height}")

# Set screen dimensions for all operations
set_screen_dimensions(screen_width, screen_height)
print(f"✓ Screen dimensions set: {SCREEN_WIDTH}x{SCREEN_HEIGHT}\n")

# Initialize OCR (lazy load)
ocr_engine = None
ocr_initialized = False

# Global region variable
CURRENT_REGION = None

def initialize_ocr():
    """Initialize EasyOCR engine"""
    global ocr_engine, ocr_initialized
    
    if ocr_initialized:
        return True
    
    if not OCR_AVAILABLE:
        print("❌ EasyOCR not available")
        return False
    
    try:
        print("🔧 Initializing OCR engine...")
        ocr_engine = easyocr.Reader(['en', 'ch_tra'], gpu=False, verbose=False)
        ocr_initialized = True
        print("✓ OCR engine ready")
        return True
    except Exception as e:
        print(f"❌ OCR init failed: {e}")
        return False


def ping_host(host, timeout=3):
    """
    Ping a host to check network connectivity
    Returns True if ping successful, False otherwise
    """
    try:
        # Use ping command (works on Windows and Linux)
        if sys.platform.startswith('win'):
            # Windows ping: -n 1 (1 packet), -w timeout in milliseconds
            result = subprocess.run(
                ['ping', '-n', '1', '-w', str(timeout * 1000), host],
                capture_output=True,
                timeout=timeout + 2
            )
        else:
            # Linux/Mac ping: -c 1 (1 packet), -W timeout in seconds
            result = subprocess.run(
                ['ping', '-c', '1', '-W', str(timeout), host],
                capture_output=True,
                timeout=timeout + 2
            )
        return result.returncode == 0
    except Exception as e:
        print(f"⚠ Ping error for {host}: {e}")
        return False


def get_phone_ip():
    """
    Get phone's IP address via ADB
    Returns IP address string or None
    """
    try:
        # Get IP address from phone
        result = d.shell("ip addr show wlan0 | grep 'inet ' | awk '{print $2}' | cut -d/ -f1")
        if result and result.strip():
            return result.strip()
        
        # Alternative method
        result = d.shell("getprop dhcp.wlan0.ipaddress")
        if result and result.strip():
            return result.strip()
        
        return None
    except Exception as e:
        print(f"⚠ Error getting phone IP: {e}")
        return None


def detect_region():
    """
    Detect current region/country by checking network connectivity to region-specific servers
    Returns: 'HK' (Hong Kong), 'MY' (Malaysia), 'SG' (Singapore), or None
    """
    global CURRENT_REGION
    
    print("\n" + "="*60)
    print("🌐 [Region Detection] Detecting current region...")
    print("="*60)
    
    # Test servers for each region
    region_servers = {
        'HK': ['8.8.8.8', '1.1.1.1'],  # Common DNS servers accessible from HK
        'MY': ['1.1.1.1', '8.8.8.8'],  # Common DNS servers accessible from MY
        'SG': ['1.1.1.1', '8.8.8.8']   # Common DNS servers accessible from SG
    }
    
    # Method 1: Check phone's network location via ADB
    try:
        # Try to get SIM card country code
        sim_country = d.shell("getprop ro.telephony.sim.country").strip()
        network_country = d.shell("getprop gsm.operator.iso-country").strip()
        
        print(f"  📱 SIM Country Code: {sim_country}")
        print(f"  📡 Network Country Code: {network_country}")
        
        if network_country:
            country_map = {
                'hk': 'HK',
                'my': 'MY',
                'sg': 'SG'
            }
            detected = country_map.get(network_country.lower())
            if detected:
                CURRENT_REGION = detected
                print(f"  ✓ Region detected via network: {detected}")
                print("="*60 + "\n")
                return detected
        
        if sim_country:
            country_map = {
                'hk': 'HK',
                'my': 'MY',
                'sg': 'SG'
            }
            detected = country_map.get(sim_country.lower())
            if detected:
                CURRENT_REGION = detected
                print(f"  ✓ Region detected via SIM: {detected}")
                print("="*60 + "\n")
                return detected
    except Exception as e:
        print(f"  ⚠ Error checking phone network info: {e}")
    
    # Method 2: Check phone's IP and geolocation (simplified)
    try:
        phone_ip = get_phone_ip()
        if phone_ip:
            print(f"  📱 Phone IP: {phone_ip}")
            # Simple heuristic: Check IP ranges (this is a simplified approach)
            # In production, you might want to use a geolocation API
    except Exception as e:
        print(f"  ⚠ Error getting phone IP: {e}")
    
    # Method 3: Ping test (fallback)
    print("  🔍 Testing network connectivity...")
    ping_results = {}
    for region, servers in region_servers.items():
        success_count = 0
        for server in servers:
            if ping_host(server, timeout=2):
                success_count += 1
        ping_results[region] = success_count > 0
    
    # If all regions can ping, we need another method
    # For now, default to HK if we can't determine
    if all(ping_results.values()):
        print("  ⚠ All regions reachable, defaulting to HK")
        CURRENT_REGION = 'HK'
    elif ping_results.get('HK'):
        CURRENT_REGION = 'HK'
        print("  ✓ Region detected: Hong Kong (HK)")
    elif ping_results.get('MY'):
        CURRENT_REGION = 'MY'
        print("  ✓ Region detected: Malaysia (MY)")
    elif ping_results.get('SG'):
        CURRENT_REGION = 'SG'
        print("  ✓ Region detected: Singapore (SG)")
    else:
        print("  ⚠ Could not determine region, defaulting to HK")
        CURRENT_REGION = 'HK'
    
    print("="*60 + "\n")
    return CURRENT_REGION


def get_price_for_region(row_data, region):
    """
    Get price based on current region
    Returns price value or None
    """
    if region == 'HK':
        return row_data.get('HKPrice', '')
    elif region == 'MY':
        return row_data.get('MYPrice', '')
    elif region == 'SG':
        return row_data.get('SGPrice', '')
    else:
        # Default to HK price
        return row_data.get('HKPrice', '')


def get_title_for_region(row_data, region):
    """
    Get title based on current region
    Returns title string
    """
    if region == 'HK':
        # Hong Kong: Use Chinese title
        return row_data.get('ProductNameCn', '')
    elif region in ['MY', 'SG']:
        # Malaysia and Singapore: Use English title
        return row_data.get('ProductNameEn', '')
    else:
        # Default to English
        return row_data.get('ProductNameEn', row_data.get('ProductNameCn', ''))


def take_screenshot():
    """Take a screenshot and convert to numpy array for OpenCV"""
    try:
        img = d.screenshot()
        img_np = np.array(img)
        # Convert RGB to BGR for OpenCV
        return cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)
    except Exception as e:
        print(f"Error taking screenshot: {e}")
        return None


def percent_to_pixels(x_pct, y_pct):
    """Convert percentage coordinates to pixel coordinates"""
    return int(x_pct * screen_width), int(y_pct * screen_height)


def click_with_jitter(x_pct, y_pct, jitter=0.02, label=""):
    """Click at percentage coordinates with random jitter using human-like click"""
    x_actual = max(0, min(1, x_pct + random.uniform(-jitter, jitter)))
    y_actual = max(0, min(1, y_pct + random.uniform(-jitter, jitter)))
    px, py = percent_to_pixels(x_actual, y_actual)
    human_click_with_pressure(px, py)
    print(f"📍 {label}: ({x_actual:.3f}, {y_actual:.3f}) -> ({px}, {py})px")
    time.sleep(WAIT_AFTER_CLICK)
    return True


def click_random_in_rect(rect, label=""):
    """Click at center point with small random offset within rectangle (offset won't exceed boundaries) using human-like click"""
    # Calculate center point
    center_x = (rect['x_min'] + rect['x_max']) / 2.0
    center_y = (rect['y_min'] + rect['y_max']) / 2.0
    
    # Calculate region size
    width = rect['x_max'] - rect['x_min']
    height = rect['y_max'] - rect['y_min']
    
    # Add small random offset (max 20% of region size, but ensure it doesn't exceed boundaries)
    max_offset_x = min(width * 0.2, (rect['x_max'] - center_x), (center_x - rect['x_min']))
    max_offset_y = min(height * 0.2, (rect['y_max'] - center_y), (center_y - rect['y_min']))
    
    # Random offset within safe range
    offset_x = random.uniform(-max_offset_x, max_offset_x)
    offset_y = random.uniform(-max_offset_y, max_offset_y)
    
    # Final coordinates (guaranteed to be within bounds)
    x_final = center_x + offset_x
    y_final = center_y + offset_y
    
    # Ensure coordinates are within bounds (safety check)
    x_final = max(rect['x_min'], min(rect['x_max'], x_final))
    y_final = max(rect['y_min'], min(rect['y_max'], y_final))
    
    px, py = percent_to_pixels(x_final, y_final)
    human_click_with_pressure(px, py)
    print(f"📍 {label} (Rect): ({x_final:.3f}, {y_final:.3f}) -> ({px}, {py})px")
    time.sleep(WAIT_AFTER_CLICK)
    return True


def find_text_with_ocr(search_text, timeout=15, fast_mode=True, region_rect=None):
    """Find text on screen using OCR with timeout
    Args:
        search_text: Text to search for
        timeout: Maximum time to wait
        fast_mode: If True, only use enhanced version (faster)
        region_rect: Optional dict with 'x_min', 'x_max', 'y_min', 'y_max' to limit search area
    """
    if not initialize_ocr():
        return None
    
    print(f"🔍 Searching for text: '{search_text}' (timeout: {timeout}s)")
    if region_rect:
        print(f"   Limited to region: ({region_rect['x_min']:.3f}, {region_rect['y_min']:.3f}) to ({region_rect['x_max']:.3f}, {region_rect['y_max']:.3f})")
    
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        screenshot = take_screenshot()
        if screenshot is None:
            time.sleep(0.5)
            continue
        
        try:
            # Crop to region if specified
            if region_rect:
                height, width = screenshot.shape[:2]
                x1 = int(region_rect['x_min'] * width)
                x2 = int(region_rect['x_max'] * width)
                y1 = int(region_rect['y_min'] * height)
                y2 = int(region_rect['y_max'] * height)
                cropped = screenshot[y1:y2, x1:x2]
            else:
                cropped = screenshot
            
            if fast_mode:
                # FAST MODE: Only use enhanced + sharpened (best for gray text)
                gray = cv2.cvtColor(cropped, cv2.COLOR_BGR2GRAY)
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                enhanced = clahe.apply(gray)
                kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
                sharpened = cv2.filter2D(enhanced, -1, kernel)
                
                result = ocr_engine.readtext(sharpened)
            else:
                # SLOW MODE: Try original image
                result = ocr_engine.readtext(cropped)
            
            # Search for matching text
            for detection in result:
                bbox = detection[0]
                text = detection[1]
                confidence = detection[2]
                
                if search_text.lower() in text.lower():
                    # Calculate center point
                    x_coords = [point[0] for point in bbox]
                    y_coords = [point[1] for point in bbox]
                    center_x = sum(x_coords) / 4
                    center_y = sum(y_coords) / 4
                    
                    # Adjust coordinates if region was cropped
                    if region_rect:
                        center_x += x1
                        center_y += y1
                    
                    # Convert to percentage
                    x_pct = center_x / screen_width
                    y_pct = center_y / screen_height
                    
                    print(f"✓ Found '{text}' at ({x_pct:.3f}, {y_pct:.3f}) [conf: {confidence:.2f}]")
                    return (x_pct, y_pct, bbox)
        
        except Exception as e:
            print(f"⚠ OCR error: {e}")
        
        time.sleep(0.5)
    
    print(f"❌ Text '{search_text}' not found within {timeout}s")
    return None


def find_image_on_screen(template_path, screenshot=None, threshold=SCREENSHOT_THRESHOLD):
    """
    Find image template on screen using OpenCV template matching
    Returns (x, y) coordinates of center if found, None otherwise
    """
    if screenshot is None:
        screenshot = take_screenshot()
    
    if screenshot is None:
        return None
    
    if not os.path.exists(template_path):
        print(f"Template image not found: {template_path}")
        return None
    
    # Load template image
    template = cv2.imread(template_path)
    if template is None:
        print(f"Failed to load template: {template_path}")
        return None
    
    # Perform template matching
    result = cv2.matchTemplate(screenshot, template, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    
    if max_val >= threshold:
        # Get template dimensions
        h, w = template.shape[:2]
        # Calculate center point
        center_x = max_loc[0] + w // 2
        center_y = max_loc[1] + h // 2
        print(f"Image found at ({center_x}, {center_y}) with confidence {max_val:.2f}")
        return (center_x, center_y)
    
    return None


def check_region_content_loaded(region_rect, timeout=10, check_interval=0.5, threshold=0.4):
    """
    Check if region has content loaded (text and icons) instead of just loading animation
    Loading state: blank area with only a red circle in center
    Loaded state: has text and icons
    
    Args:
        region_rect: Dict with 'x_min', 'x_max', 'y_min', 'y_max' (percentage coordinates)
        timeout: Maximum time to wait (seconds)
        check_interval: Time between checks (seconds)
        threshold: Threshold for content detection (default: 0.4, lower = more sensitive)
    
    Returns:
        (content_score, loaded) tuple - content score and whether content is loaded
    """
    print(f"  🔍 Checking if region has content loaded...")
    start_time = time.time()
    best_score = 0.0
    
    while time.time() - start_time < timeout:
        screenshot = take_screenshot()
        if screenshot is None:
            time.sleep(check_interval)
            continue
        
        try:
            height, width = screenshot.shape[:2]
            
            # Extract region
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                time.sleep(check_interval)
                continue
            
            # Method 1: OCR to detect text (most reliable)
            if initialize_ocr():
                gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                # Enhance for OCR
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                enhanced = clahe.apply(gray_region)
                
                result = ocr_engine.readtext(enhanced)
                
                # Count text detections with reasonable confidence
                text_count = 0
                for detection in result:
                    confidence = detection[2]
                    if confidence > 0.5:  # Reasonable confidence threshold
                        text_count += 1
                
                if text_count >= 2:  # If we find at least 2 text elements, content is loaded
                    print(f"  ✓ Found {text_count} text elements - content loaded")
                    return (text_count, True)
            
            # Method 2: Edge detection - loaded content has more edges
            gray = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            edges = cv2.Canny(gray, 50, 150)
            edge_density = np.sum(edges > 0) / (edges.shape[0] * edges.shape[1])
            
            # Method 3: Color variance - loaded content has more color variation
            # Convert to HSV and check variance
            hsv = cv2.cvtColor(region, cv2.COLOR_BGR2HSV)
            color_variance = np.var(hsv[:, :, 2])  # Variance in value channel
            
            # Method 4: Image entropy - loaded content has higher entropy
            hist = cv2.calcHist([gray], [0], None, [256], [0, 256])
            hist = hist / hist.sum()  # Normalize
            entropy = -np.sum(hist * np.log2(hist + 1e-10))  # Add small epsilon to avoid log(0)
            
            # Combined score (normalized)
            edge_score = min(edge_density * 10, 1.0)  # Normalize edge density
            color_score = min(color_variance / 1000, 1.0)  # Normalize color variance
            entropy_score = min(entropy / 8, 1.0)  # Normalize entropy (max ~8 for grayscale)
            
            combined_score = (edge_score * 0.3 + color_score * 0.3 + entropy_score * 0.4)
            best_score = max(best_score, combined_score)
            
            # Threshold: if combined score is high enough, content is loaded
            # Loading state (blank + red circle) has low scores
            # Loaded state (text + icons) has high scores
            if combined_score > threshold:  # Use configurable threshold
                print(f"  ✓ Content detected (score: {combined_score:.3f}, edges: {edge_density:.3f}, color_var: {color_variance:.1f}, entropy: {entropy:.2f})")
                return (combined_score, True)
            
        except Exception as e:
            print(f"  ⚠ Content check error: {e}")
        
        time.sleep(check_interval)
    
    print(f"  ❌ Content not loaded (best score: {best_score:.3f} < {threshold}, timeout: {timeout}s)")
    return (best_score, False)


def check_region_text_or_add(region_rect, timeout=10, check_interval=0.3):
    """
    快速检测区域是否有文字或"add"关键词，哪个更快用哪个
    用于Step 8的快速检测
    
    Args:
        region_rect: Dict with 'x_min', 'x_max', 'y_min', 'y_max' (percentage coordinates)
        timeout: Maximum time to wait (seconds)
        check_interval: Time between checks (seconds)
    
    Returns:
        (detected, method) tuple - 是否检测到内容，以及使用的检测方法
    """
    print(f"  🔍 快速检测区域文字或'add'关键词...")
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        screenshot = take_screenshot()
        if screenshot is None:
            time.sleep(check_interval)
            continue
        
        try:
            height, width = screenshot.shape[:2]
            
            # Extract region
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                time.sleep(check_interval)
                continue
            
            # Method 1: 快速OCR检测文字或"add"关键词（优先使用，更快）
            if initialize_ocr():
                gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                # Enhance for OCR
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                enhanced = clahe.apply(gray_region)
                
                result = ocr_engine.readtext(enhanced)
                
                # 检查是否有任何文字（低阈值，快速检测）
                text_found = False
                add_found = False
                for detection in result:
                    text = detection[1].upper()  # 转换为大写
                    confidence = detection[2]
                    if confidence > 0.3:  # 非常低的阈值，快速检测
                        text_found = True
                        # 检查是否包含"add"关键词
                        if 'ADD' in text:
                            add_found = True
                            print(f"  ✓ 找到'add'关键词 (置信度: {confidence:.2f})")
                            return (True, "add_keyword")
                
                if text_found:
                    print(f"  ✓ 找到文字内容")
                    return (True, "text")
            
            # Method 2: 边缘检测作为备用（如果OCR未初始化或未找到文字）
            gray = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            edges = cv2.Canny(gray, 50, 150)
            edge_density = np.sum(edges > 0) / (edges.shape[0] * edges.shape[1])
            
            # 非常低的阈值，只要有边缘就认为有内容
            if edge_density > 0.05:  # 非常低的阈值
                print(f"  ✓ 通过边缘检测找到内容 (边缘密度: {edge_density:.3f})")
                return (True, "edge")
            
        except Exception as e:
            print(f"  ⚠ 检测错误: {e}")
        
        time.sleep(check_interval)
    
    print(f"  ❌ 未检测到文字或'add' (超时: {timeout}秒)")
    return (False, None)


def click_image(image_name, timeout=DEFAULT_TIMEOUT, threshold=SCREENSHOT_THRESHOLD, required=True):
    """
    Find and click an image on screen
    Returns True if clicked, False otherwise
    """
    template_path = os.path.join(ASSETS_DIR, image_name)
    print(f"Looking for: {image_name}")
    
    start_time = time.time()
    while time.time() - start_time < timeout:
        screenshot = take_screenshot()
        if screenshot is None:
            time.sleep(0.5)
            continue
        
        coords = find_image_on_screen(template_path, screenshot, threshold)
        if coords:
            x, y = coords
            human_click_with_pressure(x, y)
            print(f"Clicked {image_name} at ({x}, {y})")
            time.sleep(WAIT_AFTER_CLICK)
            return True
        
        time.sleep(0.5)
    
    if required:
        print(f"Failed to find {image_name} within {timeout}s")
    else:
        print(f"{image_name} not found (optional)")
    
    return False


def step_delay(min_seconds=1, max_seconds=2):
    """Add random delay between steps to simulate human behavior
    Args:
        min_seconds: Minimum delay in seconds (default: 1)
        max_seconds: Maximum delay in seconds (default: 2)
    """
    delay = random.uniform(min_seconds, max_seconds)
    time.sleep(delay)


def clear_text_fast(device_serial=None):
    """
    极速清空文本框：模拟 Ctrl+A 全选 -> Delete
    """
    serial_cmd = f"-s {device_serial}" if device_serial else ""
    
    print("  🧹 正在极速清空文本...")
    
    # 1. 尝试发送 Ctrl+A (全选)
    # 注意：不同安卓版本对组合键的支持不同，下面两种方法互补
    
    # 方法 A：针对 Android 11+ 的 keycombination 命令 (最稳)
    ret = os.system(f"adb {serial_cmd} shell input keycombination 286 29")
    
    # 方法 B：针对旧版本的传统的 keyevent 连发 (备用)
    if ret != 0:
        # 按下 Ctrl，按下 A，抬起 A，抬起 Ctrl (模拟物理按键逻辑)
        # 但通常简单的 keyevent 286 29 也能生效
        os.system(f"adb {serial_cmd} shell input keyevent 286 29")

    time.sleep(0.2)  # 给系统一点反应时间选中文字
    
    # 2. 发送 Delete (删除)
    os.system(f"adb {serial_cmd} shell input keyevent 67")
    
    print("  ✓ 文本已清空")


def input_text(text, human_like=True):
    """Input text using smart hybrid method (wrapper for input_text_stealth)
    This function is kept for backward compatibility.
    All text input now uses the smart hybrid method: Native Gboard with intelligent input selection
    
    Args:
        text: Text to input (can contain Chinese characters)
        human_like: Kept for compatibility, not used (smart method is always human-like)
    """
    # Use smart hybrid input method for all text input
    input_text_stealth(text, APP_PACKAGE)


def scroll_down(duration=500):
    """Scroll down on screen"""
    print("Scrolling down...")
    try:
        # Get screen size
        info = d.window_size()
        width, height = info
        
        # Swipe from bottom to top (scroll down)
        start_x = width // 2
        start_y = int(height * 0.7)
        end_y = int(height * 0.3)
        
        d.swipe(start_x, start_y, start_x, end_y, duration / 1000.0)
        time.sleep(0.5)
    except Exception as e:
        print(f"Error scrolling: {e}")


def swipe_down_bezier(duration=600):
    """Swipe down using Bezier curve for more natural movement
    Args:
        duration: Swipe duration in milliseconds
    """
    print("  📜 Swiping down with Bezier curve...")
    try:
        # Get screen size
        width, height = d.window_size()
        
        # Start from middle-top, end at middle-bottom (swipe DOWN)
        start_x = width // 2
        start_y = int(height * 0.3)  # Start at 30% from top
        end_x = width // 2
        end_y = int(height * 0.7)    # End at 70% from top
        
        # Execute swipe (swipe from top to bottom = scroll down)
        d.swipe(start_x, start_y, end_x, end_y, duration / 1000.0)
        time.sleep(0.5)
        
    except Exception as e:
        print(f"  ⚠ Error swiping: {e}")


def swipe_up_in_region(region_rect, duration=500):
    """Swipe up within a specific region
    Args:
        region_rect: Dict with 'x_min', 'x_max', 'y_min', 'y_max'
        duration: Swipe duration in milliseconds
    """
    print(f"  📜 Swiping up in region...")
    try:
        width, height = d.window_size()
        
        # Calculate center x of region
        center_x = int((region_rect['x_min'] + region_rect['x_max']) / 2 * width)
        
        # Start from bottom of region, end at top of region
        start_y = int(region_rect['y_max'] * height)
        end_y = int(region_rect['y_min'] * height)
        
        # Execute swipe (swipe from bottom to top = scroll up)
        d.swipe(center_x, start_y, center_x, end_y, duration / 1000.0)
        time.sleep(0.5)
        
    except Exception as e:
        print(f"  ⚠ Error swiping: {e}")


def swipe_down_in_region(region_rect, duration=600, extend_distance=0.2):
    """Swipe down within a specific region with extended distance
    Args:
        region_rect: Dict with 'x_min', 'x_max', 'y_min', 'y_max'
        duration: Swipe duration in milliseconds
        extend_distance: Additional distance to swipe beyond region (as fraction of screen height)
    """
    print(f"  📜 Swiping down in region (extended distance)...")
    try:
        width, height = d.window_size()
        
        # Calculate center x of region
        center_x = int((region_rect['x_min'] + region_rect['x_max']) / 2 * width)
        
        # Start from top of region
        start_y = int(region_rect['y_min'] * height)
        # End beyond bottom of region to increase swipe distance
        end_y = int((region_rect['y_max'] + extend_distance) * height)
        # Ensure end_y doesn't exceed screen height
        end_y = min(end_y, height - 10)
        
        # Execute swipe (swipe from top to bottom = scroll down)
        d.swipe(center_x, start_y, center_x, end_y, duration / 1000.0)
        time.sleep(0.5)
        
    except Exception as e:
        print(f"  ⚠ Error swiping: {e}")


def handle_permissions():
    """Handle permission dialogs if they appear"""
    # Check for permission allow button
    if click_image("02_permission_allow.png", timeout=3, required=False):
        time.sleep(1)
    
    # Check for permission confirm button
    if click_image("03_permission_confirm.png", timeout=3, required=False):
        time.sleep(1)
    
    # Check for AI permission confirm button
    if click_image("03_permission_aiconfirm.png", timeout=3, required=False):
        time.sleep(1)


def wait_until_sell_button_ready(timeout=15, check_interval=0.5):
    """Wait until Sell button area is interactable (no intermediate clicks)
    
    Args:
        timeout: Maximum time to wait (seconds)
        check_interval: Time between checks (seconds)
    
    Returns:
        True if ready, False if timeout
    """
    print(f"⏳ Waiting for Sell button to be ready (timeout: {timeout}s)...")
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        screenshot = take_screenshot()
        if screenshot is None:
            time.sleep(check_interval)
            continue
        
        try:
            # Check if Sell button region has stable content
            # Extract Sell button region
            height, width = screenshot.shape[:2]
            y1 = int(SELL_BUTTON_RECT['y_min'] * height)
            y2 = int(SELL_BUTTON_RECT['y_max'] * height)
            x1 = int(SELL_BUTTON_RECT['x_min'] * width)
            x2 = int(SELL_BUTTON_RECT['x_max'] * width)
            
            sell_region = screenshot[y1:y2, x1:x2]
            
            # Simple check: if region is not completely black/white, assume ready
            gray = cv2.cvtColor(sell_region, cv2.COLOR_BGR2GRAY)
            mean_intensity = np.mean(gray)
            
            # If mean intensity is reasonable (not loading screen), consider ready
            if 20 < mean_intensity < 235:
                print("✓ Sell button area is ready")
                return True
        
        except Exception as e:
            print(f"⚠ Check error: {e}")
        
        time.sleep(check_interval)
    
    print(f"⚠ Timeout waiting for Sell button, proceeding anyway")
    return True  # Proceed anyway after timeout


def force_restart_app():
    """Force restart Carousell app"""
    print(f"\n🚀 Force restarting: {APP_PACKAGE}")
    try:
        # Wake up screen
        if not d.is_screen_on():
            d.keyevent("KEYCODE_WAKEUP")
            time.sleep(1)
            d.swipe(screen_width//2, int(screen_height*0.8), 
                   screen_width//2, int(screen_height*0.2), 0.5)
            time.sleep(1)
        
        # Go to home
        d.keyevent("KEYCODE_HOME")
        time.sleep(1)
        
        # Force stop
        d.shell(f"am force-stop {APP_PACKAGE}")
        time.sleep(2)
        
        # Start app
        d.shell(f"monkey -p {APP_PACKAGE} -c android.intent.category.LAUNCHER 1")
        time.sleep(3)  # Brief wait for app to start launching
        
        # Wait for Sell button to be ready (no intermediate clicks)
        wait_until_sell_button_ready(timeout=15)
        print("✓ App ready")
        
    except Exception as e:
        print(f"❌ Restart failed: {e}")


def clear_and_push_images(folder_name, max_images=9):
    """Clear phone directory and push images
    Returns: (list of successfully pushed files, count of files found on PC)
    """
    print(f"\n📷 [Image Upload] Processing folder: {folder_name}")
    
    # Step 1: Find images on PC FIRST (before pushing)
    src_path = os.path.join(PC_IMAGES_BASE, folder_name)
    
    if not os.path.exists(src_path):
        print(f"  ❌ Source folder not found: {src_path}")
        return [], 0
    
    # Step 2: Find and count images on PC
    # Use set with normalized paths to avoid duplicates (case-insensitive)
    all_files_set = set()
    image_extensions = {'.jpg', '.jpeg', '.png', '.webp'}
    
    # Iterate through all files in directory
    if os.path.exists(src_path):
        for item in Path(src_path).iterdir():
            if item.is_file():  # Only process files, not directories
                # Get extension in lowercase for case-insensitive matching
                ext = item.suffix.lower()
                if ext in image_extensions:
                    # Use normalized path (resolve symlinks, etc.) to avoid duplicates
                    normalized_path = item.resolve()
                    all_files_set.add(normalized_path)
    
    # Convert set to sorted list and limit to max_images
    all_files = sorted(list(all_files_set))[:max_images]
    
    # Debug: Print found files
    print(f"  📋 Files found ({len(all_files)} total):")
    for idx, f in enumerate(all_files, 1):
        print(f"    [{idx}] {f.name}")
    
    if not all_files:
        print("  ❌ No images found")
        return [], 0
    
    num_files_found = len(all_files)
    print(f"  📊 Found {num_files_found} images on PC")
    
    # Step 3: Clean phone directory
    print("  🧹 Cleaning phone directory...")
    d.shell(f"rm -rf {PHONE_TEMP_DIR}")
    d.shell(f"mkdir -p {PHONE_TEMP_DIR}")
    
    # Step 4: Push files and track successful pushes
    print(f"  📤 Pushing {num_files_found} images...")
    successfully_pushed = []
    
    for idx, file_path in enumerate(all_files, 1):
        dest_path = f"{PHONE_TEMP_DIR}/{file_path.name}"
        try:
            d.sync.push(str(file_path), dest_path)
            successfully_pushed.append(file_path)
            print(f"    [{idx}/{num_files_found}] {file_path.name} ✓")
        except Exception as e:
            print(f"    [{idx}/{num_files_found}] {file_path.name} ❌ {e}")
    
    # Step 5: Media scan
    print("  📡 Broadcasting media scan...")
    d.shell(f"am broadcast -a android.intent.action.MEDIA_SCANNER_SCAN_FILE -d file://{PHONE_TEMP_DIR}")
    d.shell(f"am broadcast -a android.intent.action.MEDIA_MOUNTED -d file://{PHONE_TEMP_DIR}")
    time.sleep(2)
    
    num_success = len(successfully_pushed)
    print(f"  ✓ {num_success}/{num_files_found} images successfully pushed")
    
    # Return both the file list and the count found on PC
    return successfully_pushed, num_files_found


def verify_draft_page_loaded(timeout=5, check_interval=0.5):
    """Verify if draft management page is loaded using multiple detection methods
    Uses OCR to find "Delete" or "Manage" text, and checks for grid layout
    
    Args:
        timeout: Maximum time to wait (seconds)
        check_interval: Time between checks (seconds)
    
    Returns:
        True if draft page detected, False if timeout
    """
    print(f"  🔍 Verifying draft page loaded (timeout: {timeout}s)...")
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        screenshot = take_screenshot()
        if screenshot is None:
            time.sleep(check_interval)
            continue
        
        try:
            # Method 1: Use OCR to find "Delete" text (more reliable than template matching)
            if initialize_ocr():
                # Search in bottom 30% of screen where Delete button typically appears
                height, width = screenshot.shape[:2]
                crop_y_start = int(height * 0.7)
                cropped_img = screenshot[crop_y_start:, :]
                
                # Convert to grayscale and enhance for OCR
                gray = cv2.cvtColor(cropped_img, cv2.COLOR_BGR2GRAY)
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                enhanced = clahe.apply(gray)
                
                result = ocr_engine.readtext(enhanced)
                
                # Check for "Delete" or "删除" text
                for detection in result:
                    text = detection[1].lower()
                    if 'delete' in text or '删除' in text or 'del' in text:
                        confidence = detection[2]
                        if confidence > 0.5:  # Lower threshold for text detection
                            print(f"  ✓ Found Delete text via OCR (confidence: {confidence:.2f})")
                            return True
            
            # Method 2: Check for grid layout (draft items are in a grid)
            # Draft page typically has a grid of items in the middle area
            height, width = screenshot.shape[:2]
            # Check middle area (20% to 80% of screen height) for grid-like patterns
            grid_region = screenshot[int(height*0.2):int(height*0.8), :]
            gray = cv2.cvtColor(grid_region, cv2.COLOR_BGR2GRAY)
            
            # Use edge detection to find grid patterns
            edges = cv2.Canny(gray, 50, 150)
            # Count edge pixels - grid layouts have more edges
            edge_density = np.sum(edges > 0) / (edges.shape[0] * edges.shape[1])
            
            # Draft pages typically have higher edge density due to grid layout
            if edge_density > 0.15:  # Threshold for grid-like layout
                print(f"  ✓ Detected grid layout (edge density: {edge_density:.3f})")
                return True
            
            # Method 3: Check if Manage button area changed (clicked successfully)
            # If we clicked Manage and page changed, the top area should be different
            # This is a fallback - check if top-right area has changed
            
        except Exception as e:
            print(f"  ⚠ Detection error: {e}")
        
        time.sleep(check_interval)
    
    print(f"  ❌ Draft page not detected within {timeout}s")
    return False


def clear_drafts():
    """Clear draft listings if they exist
    Returns: 'drafts_cleared' if drafts were deleted, 'no_drafts' if no drafts found
    Strategy: Click Manage -> Verify Delete button -> Branch accordingly
    """
    print("\n🧹 [Draft Cleanup] Checking for drafts...")
    
    # Step 1: Click Manage button area
    print("  🔍 Clicking Manage button...")
    click_random_in_rect(MANAGE_BUTTON, "Manage Button")
    
    # Step 2: Verify draft page loaded (check for Delete button template)
    if verify_draft_page_loaded(timeout=5, check_interval=0.5):
        print("  🚨 Drafts detected! Entering cleanup mode...")
        
        # Step 3: Select draft items (avoid Delete button zone)
        print("  📄 Selecting draft items...")
        delete_zone_y_min = DELETE_BUTTON['y_min']
        
        selected_count = 0
        for idx, pos in enumerate(DRAFT_GRID['positions'], 1):
            # Check if this position is above the delete button
            if pos['y_max'] < delete_zone_y_min:
                print(f"    Selecting draft {idx}...")
                click_random_in_rect(pos, f"Draft {idx}")
                time.sleep(0.5)
                selected_count += 1
            else:
                print(f"    Skipping position {idx} (too close to Delete button)")
        
        if selected_count == 0:
            print("  ⚠ No safe items to select")
            d.keyevent("KEYCODE_BACK")  # Go back
            time.sleep(1)
            return 'no_drafts'
        
        # Step 4: Click Delete button
        print(f"  🗑️ Deleting {selected_count} drafts...")
        click_random_in_rect(DELETE_BUTTON, "Delete Button")
        time.sleep(2)
        
        # Step 5: Go back to main screen
        d.keyevent("KEYCODE_BACK")
        time.sleep(1)
        
        print("  ✓ Draft cleanup complete")
        return 'drafts_cleared'
    else:
        # Delete button not found -> No drafts or Manage button not active
        print("  ✓ No drafts found (Delete button not detected)")
        print("  ℹ️ Already in gallery view, proceeding to image selection...")
        # CRUCIAL: Do NOT press Back - we're already in the right place
        return 'no_drafts'


def select_images(num_images):
    """Select images from gallery using hardcoded coordinates
    Args:
        num_images: Number of images to select
    Returns:
        True if successful
    """
    print(f"\n�️ [Image Selection] Selecting {num_images} images...")
    
    # Step A: Open Folder Dropdown
    print("  📂 Opening folder dropdown...")
    folder_dropdown_rect = {
        'x_min': 0.072,
        'x_max': 0.353,
        'y_min': 0.112,
        'y_max': 0.135
    }
    click_random_in_rect(folder_dropdown_rect, "Folder Dropdown")
    time.sleep(1.5)
    
    # Step B: Select Target Folder (Album_Cache_Temp)
    print("  📁 Selecting Album_Cache_Temp folder...")
    # Use coordinates provided by user: (0.375, 0.201) and (0.114, 0.185)
    # Calculate rectangle bounds and center point with random offset
    x_coords = [0.375, 0.114]
    y_coords = [0.201, 0.185]
    
    album_folder_rect = {
        'x_min': min(x_coords),
        'x_max': max(x_coords),
        'y_min': min(y_coords),
        'y_max': max(y_coords)
    }
    
    # Calculate center point
    center_x = (album_folder_rect['x_min'] + album_folder_rect['x_max']) / 2
    center_y = (album_folder_rect['y_min'] + album_folder_rect['y_max']) / 2
    
    # Calculate width and height for random offset
    rect_width = album_folder_rect['x_max'] - album_folder_rect['x_min']
    rect_height = album_folder_rect['y_max'] - album_folder_rect['y_min']
    
    # Add random offset within the rectangle (use 30% of width/height as max offset)
    offset_x = random.uniform(-rect_width * 0.15, rect_width * 0.15)
    offset_y = random.uniform(-rect_height * 0.15, rect_height * 0.15)
    
    final_x = max(album_folder_rect['x_min'], min(album_folder_rect['x_max'], center_x + offset_x))
    final_y = max(album_folder_rect['y_min'], min(album_folder_rect['y_max'], center_y + offset_y))
    
    click_with_jitter(final_x, final_y, jitter=0.01, label="Album_Cache_Temp")
    time.sleep(2.5)  # Wait for images to load
    
    # Step C: Select Images from Grid
    print(f"  🎯 Selecting {num_images} images from grid...")
    
    # Grid configuration
    grid_config = {
        'top_left': (0.033, 0.198),
        'bottom_right': (1.012, 0.940),
        'columns': 3,
        'rows': 5,
        'skip_first': True  # Skip camera button
    }
    
    # Calculate grid cell dimensions
    x1, y1 = grid_config['top_left']
    x2, y2 = grid_config['bottom_right']
    cols = grid_config['columns']
    rows = grid_config['rows']
    
    cell_w = (x2 - x1) / cols
    cell_h = (y2 - y1) / rows
    
    # Select images (skip first slot which is camera)
    # IMPORTANT: Only select the number of images that were actually pushed
    for i in range(num_images):
        # Real index (skip camera at position 0)
        real_idx = i + 1 if grid_config['skip_first'] else i
        
        # Calculate grid position
        col = real_idx % cols
        row = real_idx // cols
        
        # Calculate center point of cell
        cx = x1 + (col + 0.5) * cell_w
        cy = y1 + (row + 0.5) * cell_h
        
        # Click with small jitter
        click_with_jitter(cx, cy, jitter=0.03, label=f"Image {i+1}/{num_images}")
        time.sleep(0.8)
    
    # Step D: Click Next Button
    print("  ➡️ Clicking Next button...")
    # Click region strictly within the four coordinates provided:
    # (0.094, 0.920), (0.906, 0.917), (0.914, 0.956), (0.081, 0.955)
    # Calculate bounding rectangle
    next_button_rect = {
        'x_min': min(0.094, 0.906, 0.914, 0.081),  # 0.081
        'x_max': max(0.094, 0.906, 0.914, 0.081),  # 0.914
        'y_min': min(0.920, 0.917, 0.956, 0.955),  # 0.917
        'y_max': max(0.920, 0.917, 0.956, 0.955)   # 0.956
    }
    # Click within this strict region
    click_random_in_rect(next_button_rect, "Next Button")
    time.sleep(3)
    
    print("  ✓ Image selection complete")
    return True


def click_grid_image(idx):
    """Click image in grid (auto-skip camera at index 0)"""
    cols = IMAGE_GRID['columns']
    rows = IMAGE_GRID['rows']
    
    # Real index (skip camera if needed)
    real_idx = idx + 1 if IMAGE_GRID['skip_first'] else idx
    
    col = real_idx % cols
    row = real_idx // cols
    
    # Calculate cell bounds
    x1, y1 = IMAGE_GRID['top_left']
    x2, y2 = IMAGE_GRID['bottom_right']
    
    cell_w = (x2 - x1) / cols
    cell_h = (y2 - y1) / rows
    
    # Center point
    cx = x1 + (col + 0.5) * cell_w
    cy = y1 + (row + 0.5) * cell_h
    
    click_with_jitter(cx, cy, jitter=IMAGE_GRID['cell_jitter'], 
                     label=f"Image {idx+1}")
    time.sleep(0.8)


def upload_images_phase(row_data):
    """
    Phase 0: Upload images to Carousell
    Returns True if successful
    """
    print("\n" + "="*60)
    print("PHASE 0: IMAGE UPLOAD")
    print("="*60)
    
    # Get folder name from Excel Column J (index 9)
    folder_name = str(row_data.get('ImageFolder', '')).strip()
    
    if not folder_name or folder_name == 'nan':
        print("❌ No image folder specified in Excel")
        return False
    
    # Step 1: Push images and get count
    files, num_files_found = clear_and_push_images(folder_name)
    if not files:
        print("❌ No images were successfully pushed")
        return False
    
    # Store the number of files found on PC (before pushing)
    # This will be used for image selection
    num_images_to_select = num_files_found
    print(f"  📌 Will select {num_images_to_select} images based on files found on PC")
    
    # Step 2: Restart app
    force_restart_app()
    
    # Step 3: Click Sell button to enter gallery
    print("\n📍 [Step 3] Clicking Sell button...")
    click_random_in_rect(SELL_BUTTON_RECT, "Sell Button")
    time.sleep(4)  # Wait for gallery to load
    
    # Step 4: Clear drafts with proper branching
    draft_status = clear_drafts()
    
    # Step 5: Select images (workflow continues regardless of draft status)
    if draft_status == 'drafts_cleared':
        # Drafts were cleared, we're back at main screen
        # Need to click Sell button again
        print("\n📍 [Step 5] Re-entering gallery after draft cleanup...")
        click_random_in_rect(SELL_BUTTON_RECT, "Sell Button")
        time.sleep(4)
    elif draft_status == 'no_drafts':
        # No drafts found, already in gallery view
        print("\n� [Step 5] Continuing in gallery view...")
    
    # Step 6: Select images using hardcoded coordinates
    # Use the count found on PC (before pushing) to determine how many to select
    if not select_images(num_images_to_select):
        print("⚠ Image selection failed, aborting")
        return False
    
    # Step delay between Step 6 and Step 8 (Step 7)
    step_delay(1, 2)
    
    # Step 8: 在Step 7点击之后等待0.6秒，然后检测区域是否有文字或"add"
    print("\n⏳ [Step 8] Waiting 0.6 seconds after Step 7, then checking for text or 'add'...")
    time.sleep(0.6)  # 固定等待0.6秒
    
    # Detection region: (0.200, 0.351) to (0.781, 0.564)
    detection_region = {
        'x_min': 0.200,
        'x_max': 0.781,
        'y_min': 0.351,
        'y_max': 0.564
    }
    
    # 使用快速检测函数，检测文字或"add"，哪个更快用哪个
    print("  🔍 快速检测区域文字或'add'关键词（低阈值，快速检测）...")
    detected, method = check_region_text_or_add(
        detection_region,
        timeout=10,  # 10秒超时
        check_interval=0.3  # 每0.3秒检测一次
    )
    
    # 如果检测到内容，点击指定区域
    if detected:
        print(f"  ✓ 检测到内容 (方法: {method}) - 点击目标区域")
        # Click region: (0.353, 0.542) to (0.622, 0.569)
        click_region = {
            'x_min': 0.353,
            'x_max': 0.622,
            'y_min': 0.542,
            'y_max': 0.569
        }
        click_random_in_rect(click_region, "检测到内容后点击区域")
    else:
        print(f"  ⚠ 未检测到内容 - 继续执行")
    
    # Step delay between Step 8 and Step 9
    step_delay(1, 2)
    
    # Step 9: Swipe down and click Category input (with retry logic)
    def execute_step9():
        """Execute Step 9: Swipe down and click Category input"""
        print("\n📍 [Step 9] Swiping down and clicking Category input...")
        
        # First swipe down using Bezier curve
        swipe_down_bezier(duration=600)
        time.sleep(1)
        
        # Second swipe down (additional swipe as requested)
        print("  📜 Second swipe down...")
        swipe_down_bezier(duration=600)
        time.sleep(1)
        
        # Click Category input area directly using provided coordinates
        # Coordinates: (0.069, 0.499) to (0.781, 0.542)
        category_input_rect = {
            'x_min': 0.069,
            'x_max': 0.781,
            'y_min': 0.499,
            'y_max': 0.542
        }
        print("  🎯 Clicking Category input area...")
        click_random_in_rect(category_input_rect, "Category Input")
        time.sleep(1)
        return True
    
    # Execute Step 9 first time
    execute_step9()
    
    # Step 9.1: Wait 6-9 seconds, then check if content is loaded (to verify network status)
    # Region to check: (0.039, 0.412) to (0.992, 0.969)
    content_check_region = {
        'x_min': 0.039,
        'x_max': 0.992,
        'y_min': 0.412,
        'y_max': 0.969
    }
    
    def check_content_after_wait():
        """Wait 6-9 seconds, then check if content is loaded"""
        # Wait 6-9 seconds before checking (random delay)
        wait_time = random.uniform(6, 9)
        print(f"\n⏳ [Step 9.1] Waiting {wait_time:.1f} seconds before content check...")
        time.sleep(wait_time)
        
        # Now start checking if content is loaded (with reasonable timeout for detection)
        print(f"🔍 Starting content check (to verify network status and page load)...")
        content_score, content_loaded = check_region_content_loaded(
            content_check_region,
            timeout=10,  # Give enough time for detection
            check_interval=0.5
        )
        return content_score, content_loaded
    
    # First attempt: wait and check
    content_score, content_loaded = check_content_after_wait()
    
    if not content_loaded:
        # First attempt failed, click retry area and retry Step 9
        print(f"\n⚠️ Content check failed (score: {content_score:.3f}), page may not be loaded")
        print("  🔄 Attempting retry...")
        
        # Click retry area: (0.486, 0.105) to (0.950, 0.181)
        retry_rect = {
            'x_min': 0.486,
            'x_max': 0.950,
            'y_min': 0.105,
            'y_max': 0.181
        }
        print("  🎯 Clicking retry area...")
        click_random_in_rect(retry_rect, "Retry Area")
        time.sleep(2)
        
        # Retry Step 9
        print("\n🔄 Retrying Step 9...")
        execute_step9()
        
        # Wait 6-9 seconds again, then check content again
        content_score, content_loaded = check_content_after_wait()
        
        if not content_loaded:
            print(f"\n❌ Content check failed again (score: {content_score:.3f})")
            print("  ⚠️ Page load verification failed - skipping this account and moving to next one...")
            return False
    
    print(f"  ✓ Content check passed (score: {content_score:.3f}) - page loaded successfully")
    
    # Step 10: 在Step 9.1判断有文字出现之后就点击输入区域，然后输入"others"
    print("\n⌨️ [Step 10] Clicking input area and typing 'others'...")
    
    # 点击输入区域: (0.044, 0.430) 到 (0.550, 0.479)
    input_click_rect = {
        'x_min': 0.044,
        'x_max': 0.550,
        'y_min': 0.430,
        'y_max': 0.479
    }
    print("  🎯 Clicking input area...")
    click_random_in_rect(input_click_rect, "Input Area")
    time.sleep(1)
    
    # 输入 "others" - 使用智能混合输入方式
    print("  ⌨️ Typing 'others'...")
    input_text_stealth("others", APP_PACKAGE)
    
    # Step 10.1: 输入完成后点击指定区域
    # 点击区域: (0.136, 0.961) 到 (0.206, 0.965) 到 (0.200, 0.984) 到 (0.114, 0.988)
    print("  🎯 Clicking post-input region...")
    post_input_click_rect = {
        'x_min': min(0.136, 0.206, 0.200, 0.114),  # 0.114
        'x_max': max(0.136, 0.206, 0.200, 0.114),  # 0.206
        'y_min': min(0.961, 0.965, 0.984, 0.988),  # 0.961
        'y_max': max(0.961, 0.965, 0.984, 0.988)   # 0.988
    }
    click_random_in_rect(post_input_click_rect, "Post-Input Click Region")
    time.sleep(1)
    
    # Wait 1-2 seconds after input
    step_delay(1, 2)
    
    # Step delay between Step 10 and Step 11
    step_delay(1, 2)
    
    # Step 11: Select "Others in Services" from dropdown
    print("\n📍 [Step 11] Selecting 'Others in Services'...")
    
    # Step 11.2: Click fixed selection area
    # Click region: (0.031, 0.598) to (0.369, 0.634)
    selection_rect = {
        'x_min': 0.031,
        'x_max': 0.369,
        'y_min': 0.598,
        'y_max': 0.634
    }
    print("  🎯 Clicking selection area...")
    click_random_in_rect(selection_rect, "Others in Services Selection")
    
    # Step 11.2.1: Click additional region after selection
    # Click region: (0.633, 0.223), (0.836, 0.240), (0.833, 0.217), (0.656, 0.247)
    additional_click_region = {
        'x_min': min(0.633, 0.836, 0.833, 0.656),  # 0.633
        'x_max': max(0.633, 0.836, 0.833, 0.656),  # 0.836
        'y_min': min(0.223, 0.240, 0.217, 0.247),  # 0.217
        'y_max': max(0.223, 0.240, 0.217, 0.247)   # 0.247
    }
    print("  🎯 Clicking additional region...")
    click_random_in_rect(additional_click_region, "Additional Click After Selection")
    
    # Wait 2-3 seconds after additional click
    wait_time = random.uniform(2, 3)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds after additional click...")
    time.sleep(wait_time)
    
    # Step 11.2.2: Click second region
    # Click region: (0.636, 0.565), (0.800, 0.573), (0.806, 0.595), (0.667, 0.598)
    second_click_region = {
        'x_min': min(0.636, 0.800, 0.806, 0.667),  # 0.636
        'x_max': max(0.636, 0.800, 0.806, 0.667),  # 0.806
        'y_min': min(0.565, 0.573, 0.595, 0.598),  # 0.565
        'y_max': max(0.565, 0.573, 0.595, 0.598)   # 0.598
    }
    print("  🎯 Clicking second region...")
    click_random_in_rect(second_click_region, "Second Click Region")
    
    # Wait 3-4 seconds after second click
    wait_time = random.uniform(3, 4)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds after second click...")
    time.sleep(wait_time)
    
    # Step delay between Step 11.2.2 and Step 11.3
    step_delay(1, 2)
    
    # Step 11.3: Wait 2-3 seconds (reduced from 8-10 seconds)
    wait_time = random.uniform(2, 3)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step 11.4: Swipe down (no additional wait after swipe)
    print("  📜 Swiping down...")
    swipe_down_bezier(duration=600)
    
    # Step delay between Step 11.4 and Step 11.5
    step_delay(1, 2)
    
    # Step 11.5: Click title input area (0.061, 0.596) to (0.431, 0.625)
    title_input_rect = {
        'x_min': 0.061,
        'x_max': 0.431,
        'y_min': 0.596,
        'y_max': 0.625
    }
    print("  🎯 Clicking title input area...")
    click_random_in_rect(title_input_rect, "Title Input")
    
    # Step delay between Step 11.5 and Step 11.6
    step_delay(1, 2)
    
    # Step 11.6: Input ProductNameCn (Excel C column)
    product_name_cn = row_data.get('ProductNameCn', '')
    print(f"  ⌨️ Inputting ProductNameCn (Excel C column): {product_name_cn}")
    input_text_stealth(product_name_cn, APP_PACKAGE)
    
    # Step 11.6.1: Click first region after input
    # Click region: 左上(128, 2275) 右下(212, 2376)
    print("  🎯 [Step 11.6.1] Clicking first region after ProductNameCn input...")
    step11_6_1_region = pixels_to_rect(128, 2275, 212, 2376)
    click_random_in_rect(step11_6_1_region, "Step 11.6.1 Region")
    time.sleep(1)
    
    # Step 11.6.2: Click second region
    # Click region: 左上(62, 1667) 右下(195, 1728)
    print("  🎯 [Step 11.6.2] Clicking second region...")
    step11_6_2_region = pixels_to_rect(62, 1667, 195, 1728)
    click_random_in_rect(step11_6_2_region, "Step 11.6.2 Region")
    time.sleep(1)
    
    # Step 11.6.3: Input Excel N column content (Description)
    print("  ⌨️ [Step 11.6.3] Inputting Excel N column content (Description)...")
    description_content = row_data.get('Description', '')
    if not description_content or pd.isna(description_content):
        # Fallback to ProductNameEn if N column is empty
        description_content = row_data.get('ProductNameEn', '')
        print(f"  ⌨️ Inputting ProductNameEn as description (N column empty): {description_content}")
    else:
        print(f"  ⌨️ Inputting Excel N column content as description: {description_content}")
    input_text_stealth(str(description_content), APP_PACKAGE)
    time.sleep(1)
    
    # Step delay between Step 11.6 and Step 14
    step_delay(1, 2)
    
    # Step 14: Click price input area and input HK price
    print("\n📍 [Step 14] Clicking price input area...")
    # Click region: 左上(182, 1238) 右下(597, 1317)
    # Convert pixel coordinates to percentage using actual device resolution
    price_input_region = pixels_to_rect(182, 1238, 597, 1317)
    click_random_in_rect(price_input_region, "Price Input")
    time.sleep(1)
    
    # Clear existing text in price input before inputting
    # Use native clear text method (Ctrl+A -> Delete)
    print("  🧹 Clearing existing text in price input (if any)...")
    clear_text_native()
    
    # Input HK price
    hk_price = row_data.get('HKPrice', '')
    print(f"  ⌨️ Inputting HK price: {hk_price}")
    input_text_stealth(str(hk_price), APP_PACKAGE)
    time.sleep(1)
    
    # Step 14.1: First large swipe down (80% of screen height)
    # Swipe region: 左上(60, 516) 右下(1007, 1442)
    print("\n📍 [Step 14.1] First large swipe down (80% of screen height)...")
    swipe_down_large_distance_in_pixel_region(60, 516, 1007, 1442, swipe_distance_ratio=0.8, duration=600)
    time.sleep(1)
    
    # Step 14.2: Second large swipe down (80% of screen height, different position)
    # Swipe region: 左上(60, 516) 右下(1007, 1442) (same region, but different start position)
    print("\n📍 [Step 14.2] Second large swipe down (80% of screen height, different position)...")
    swipe_down_large_distance_in_pixel_region(60, 516, 1007, 1442, swipe_distance_ratio=0.8, duration=600)
    time.sleep(1)
    
    # Step 14.5: Click region before Step 15 (replaced swipe down)
    print("\n📍 [Step 14.5] Clicking region before Step 15...")
    # Click region: 左上(77, 1680) 右下(189, 1749)
    # Convert pixel coordinates to percentage using actual device resolution
    step14_5_click_region = pixels_to_rect(77, 1680, 189, 1749)
    click_random_in_rect(step14_5_click_region, "Step 14.5 Click Region")
    time.sleep(0.5)
    
    # Step 15: Click submit/next button
    print("\n📍 [Step 15] Clicking submit/next button...")
    # Click region: (0.125, 0.906), (0.894, 0.915), (0.903, 0.956), (0.150, 0.951)
    submit_button_region = {
        'x_min': min(0.125, 0.894, 0.903, 0.150),  # 0.125
        'x_max': max(0.125, 0.894, 0.903, 0.150),  # 0.903
        'y_min': min(0.906, 0.915, 0.956, 0.951),  # 0.906
        'y_max': max(0.906, 0.915, 0.956, 0.951)   # 0.956
    }
    click_random_in_rect(submit_button_region, "Submit/Next Button")
    
    # Wait 2-3 seconds after clicking
    wait_time = random.uniform(2, 3)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step 16: Click delivery/meetup option
    print("\n📍 [Step 16] Clicking delivery/meetup option...")
    # Click region: (0.269, 0.816), (0.603, 0.816), (0.614, 0.880), (0.319, 0.875)
    delivery_option_region = {
        'x_min': min(0.269, 0.603, 0.614, 0.319),  # 0.269
        'x_max': max(0.269, 0.603, 0.614, 0.319),  # 0.614
        'y_min': min(0.816, 0.816, 0.880, 0.875),  # 0.816
        'y_max': max(0.816, 0.816, 0.880, 0.875)   # 0.880
    }
    click_random_in_rect(delivery_option_region, "Delivery/Meetup Option")
    time.sleep(1)
    
    # Step 17: Click final submit button and wait for blue button
    print("\n📍 [Step 17] Clicking final submit button...")
    # Click region: (0.125, 0.906), (0.894, 0.915), (0.903, 0.956), (0.150, 0.951)
    final_submit_region = {
        'x_min': min(0.125, 0.894, 0.903, 0.150),  # 0.125
        'x_max': max(0.125, 0.894, 0.903, 0.150),  # 0.903
        'y_min': min(0.906, 0.915, 0.956, 0.951),  # 0.906
        'y_max': max(0.906, 0.915, 0.956, 0.951)   # 0.956
    }
    click_random_in_rect(final_submit_region, "Final Submit Button")
    
    # Wait 13-15 seconds before checking for "Other" text
    wait_time = random.uniform(13, 15)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before checking for 'Other' text...")
    time.sleep(wait_time)
    
    # Step 17.1: Check for "Other" text in specified region (replaced blue button detection)
    def check_other_text(region_rect, timeout=10, check_interval=0.5):
        """Check if "Other" text appears in region
        Args:
            region_rect: Dict with 'x_min', 'x_max', 'y_min', 'y_max'
            timeout: Maximum time to wait (seconds)
            check_interval: Time between checks (seconds)
        Returns:
            True if "Other" text found, False if timeout
        """
        print(f"  🔍 Checking for 'Other' text (timeout: {timeout}s)...")
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            screenshot = take_screenshot()
            if screenshot is None:
                time.sleep(check_interval)
                continue
            
            try:
                height, width = screenshot.shape[:2]
                
                # Extract region
                x1 = int(region_rect['x_min'] * width)
                x2 = int(region_rect['x_max'] * width)
                y1 = int(region_rect['y_min'] * height)
                y2 = int(region_rect['y_max'] * height)
                
                region = screenshot[y1:y2, x1:x2]
                
                if region.size == 0:
                    time.sleep(check_interval)
                    continue
                
                # Use OCR to find "Other" text
                if initialize_ocr():
                    gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                    # Enhance for OCR
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                    enhanced = clahe.apply(gray_region)
                    
                    result = ocr_engine.readtext(enhanced)
                    
                    # Search for "Other" text (case-insensitive)
                    for detection in result:
                        text = detection[1].lower()
                        confidence = detection[2]
                        # Check if text contains "other" (to match "Other", "Othertools", etc.)
                        if 'other' in text and confidence > 0.5:
                            print(f"  ✓ 'Other' text detected: '{detection[1]}' (confidence: {confidence:.3f})")
                            return True
                
            except Exception as e:
                print(f"  ⚠ 'Other' text check error: {e}")
            
            time.sleep(check_interval)
        
        print(f"  ❌ 'Other' text not found within {timeout}s")
        return False
    
    # "Other" text detection region: (0.031, 0.450) to (0.036, 0.522) to (0.950, 0.517) to (0.939, 0.451)
    other_text_detection_region = {
        'x_min': min(0.031, 0.036, 0.950, 0.939),  # 0.031
        'x_max': max(0.031, 0.036, 0.950, 0.939),  # 0.950
        'y_min': min(0.450, 0.522, 0.517, 0.451),  # 0.450
        'y_max': max(0.450, 0.522, 0.517, 0.451)   # 0.522
    }
    
    # Blue button click region: (0.358, 0.279) to (0.367, 0.316) to (0.644, 0.281) to (0.656, 0.314)
    # Keep the same click region as before (user said don't change the flow)
    blue_button_click_region = {
        'x_min': min(0.358, 0.367, 0.644, 0.656),  # 0.358
        'x_max': max(0.358, 0.367, 0.644, 0.656),  # 0.656
        'y_min': min(0.279, 0.316, 0.281, 0.314),  # 0.279
        'y_max': max(0.279, 0.316, 0.281, 0.314)   # 0.316
    }
    
    # First attempt: check for "Other" text
    other_text_found = check_other_text(other_text_detection_region, timeout=10, check_interval=0.5)
    
    if other_text_found:
        print("  ✓ 'Other' text found, clicking blue button region...")
        click_random_in_rect(blue_button_click_region, "Blue Button")
    else:
        # Retry Step 17 if "Other" text not found
        print("  ⚠ 'Other' text not found, retrying Step 17...")
        # Click final submit button again
        click_random_in_rect(final_submit_region, "Final Submit Button (Retry)")
        
        # Wait 13-15 seconds again
        wait_time = random.uniform(13, 15)
        print(f"  ⏳ Waiting {wait_time:.1f} seconds before checking again...")
        time.sleep(wait_time)
        
        # Check for "Other" text again
        other_text_found = check_other_text(other_text_detection_region, timeout=10, check_interval=0.5)
        
        if other_text_found:
            print("  ✓ 'Other' text found on retry, clicking blue button region...")
            click_random_in_rect(blue_button_click_region, "Blue Button (Retry)")
        else:
            print("  ⚠ 'Other' text still not found after retry, continuing anyway...")
    
    # Step delay between Step 17 and Step 18
    step_delay(1, 2)
    
    # Step 18: Click region (0.444, 0.274), (0.544, 0.273), (0.547, 0.286), (0.464, 0.284)
    print("\n📍 [Step 18] Clicking region...")
    step18_region = {
        'x_min': min(0.444, 0.544, 0.547, 0.464),  # 0.444
        'x_max': max(0.444, 0.544, 0.547, 0.464),  # 0.547
        'y_min': min(0.274, 0.273, 0.286, 0.284),  # 0.273
        'y_max': max(0.274, 0.273, 0.286, 0.284)   # 0.286
    }
    click_random_in_rect(step18_region, "Step 18 Region")
    time.sleep(1)
    
    # Step delay between Step 18 and Step 19
    step_delay(1, 2)
    
    # Step 19: Click region (0.053, 0.329), (0.456, 0.324), (0.453, 0.505), (0.078, 0.504)
    print("\n📍 [Step 19] Clicking region...")
    step19_region = {
        'x_min': min(0.053, 0.456, 0.453, 0.078),  # 0.053
        'x_max': max(0.053, 0.456, 0.453, 0.078),  # 0.456
        'y_min': min(0.329, 0.324, 0.505, 0.504),  # 0.324
        'y_max': max(0.329, 0.324, 0.505, 0.504)   # 0.505
    }
    click_random_in_rect(step19_region, "Step 19 Region")
    time.sleep(1)
    
    # Step delay between Step 19 and Step 20
    step_delay(1, 2)
    
    # Step 20: Wait 5-6 seconds, then click two regions
    print("\n📍 [Step 20] Waiting 5-6 seconds, then clicking regions...")
    wait_time = random.uniform(5, 6)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step 20.1: Click first region (0.925, 0.068), (0.981, 0.068), (0.975, 0.087), (0.925, 0.089)
    step20_1_region = {
        'x_min': min(0.925, 0.981, 0.975, 0.925),  # 0.925
        'x_max': max(0.925, 0.981, 0.975, 0.925),  # 0.981
        'y_min': min(0.068, 0.068, 0.087, 0.089),  # 0.068
        'y_max': max(0.068, 0.068, 0.087, 0.089)   # 0.089
    }
    print("  🎯 Clicking first region...")
    click_random_in_rect(step20_1_region, "Step 20.1 Region")
    time.sleep(1)
    
    # Step 20.2: Click second region (0.531, 0.125), (0.725, 0.125), (0.725, 0.141), (0.536, 0.144)
    step20_2_region = {
        'x_min': min(0.531, 0.725, 0.725, 0.536),  # 0.531
        'x_max': max(0.531, 0.725, 0.725, 0.536),  # 0.725
        'y_min': min(0.125, 0.125, 0.141, 0.144),  # 0.125
        'y_max': max(0.125, 0.125, 0.141, 0.144)   # 0.144
    }
    print("  🎯 Clicking second region...")
    click_random_in_rect(step20_2_region, "Step 20.2 Region")
    
    # Step 20.3: Wait 10-12 seconds, then check if region has content (text and images)
    print("\n📍 [Step 20.3] Waiting 10-12 seconds, then checking for content...")
    wait_time = random.uniform(10, 12)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Content check region: (0.033, 0.311) to (0.942, 0.509)
    content_check_region_20_3 = {
        'x_min': min(0.033, 0.961, 0.942, 0.039),  # 0.033
        'x_max': max(0.033, 0.961, 0.942, 0.039),  # 0.961
        'y_min': min(0.311, 0.324, 0.509, 0.510),  # 0.311
        'y_max': max(0.311, 0.324, 0.509, 0.510)   # 0.510
    }
    
    # Check if content is loaded (text and images)
    print(f"  🔍 Checking if region has content (text and images)...")
    content_score, content_loaded = check_region_content_loaded(
        content_check_region_20_3,
        timeout=5,
        check_interval=0.5,
        threshold=0.3  # Lower threshold for faster detection
    )
    
    if not content_loaded:
        # Retry: Repeat Step 20.1 and 20.2
        print(f"  ⚠ Content check failed (score: {content_score:.3f}), retrying Step 20.1 and 20.2...")
        
        # Repeat Step 20.1
        print("  🔄 Repeating Step 20.1...")
        click_random_in_rect(step20_1_region, "Step 20.1 Region (Retry)")
        time.sleep(1)
        
        # Repeat Step 20.2
        print("  🔄 Repeating Step 20.2...")
        click_random_in_rect(step20_2_region, "Step 20.2 Region (Retry)")
        
        # Wait 10-12 seconds again
        wait_time = random.uniform(10, 12)
        print(f"  ⏳ Waiting {wait_time:.1f} seconds before checking again...")
        time.sleep(wait_time)
        
        # Check content again
        print(f"  🔍 Checking content again...")
        content_score, content_loaded = check_region_content_loaded(
            content_check_region_20_3,
            timeout=5,
            check_interval=0.5,
            threshold=0.3
        )
        
        if not content_loaded:
            print(f"  ❌ Content check failed again (score: {content_score:.3f}), canceling task...")
            return False
    
    print(f"  ✓ Content check passed (score: {content_score:.3f}) - proceeding to next step")
    time.sleep(1)
    
    # Step delay between Step 20 and Step 21
    step_delay(1, 2)
    
    # Step 21: Click region (0.072, 0.395), (0.622, 0.398), (0.614, 0.421), (0.083, 0.422)
    print("\n📍 [Step 21] Clicking region...")
    step21_region = {
        'x_min': min(0.072, 0.622, 0.614, 0.083),  # 0.072
        'x_max': max(0.072, 0.622, 0.614, 0.083),  # 0.622
        'y_min': min(0.395, 0.398, 0.421, 0.422),  # 0.395
        'y_max': max(0.395, 0.398, 0.421, 0.422)   # 0.422
    }
    click_random_in_rect(step21_region, "Step 21 Region")
    time.sleep(1)
    
    # Step delay between Step 21 and Step 22
    step_delay(1, 2)
    
    # Step 22: Use Step 9.1 content check logic, then input "sneakers"
    print("\n📍 [Step 22] Checking content (using Step 9.1 logic), then inputting 'sneakers'...")
    
    # Use the same content check region as Step 9.1: (0.039, 0.412) to (0.992, 0.969)
    content_check_region_step22 = {
        'x_min': 0.039,
        'x_max': 0.992,
        'y_min': 0.412,
        'y_max': 0.969
    }
    
    # Wait 6-9 seconds before checking (same as Step 9.1)
    wait_time = random.uniform(6, 9)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before content check...")
    time.sleep(wait_time)
    
    # Check if content is loaded (using same logic as Step 9.1)
    print(f"  🔍 Starting content check (to verify network status and page load)...")
    content_score, content_loaded = check_region_content_loaded(
        content_check_region_step22,
        timeout=10,
        check_interval=0.5
    )
    
    if content_loaded:
        print(f"  ✓ Content check passed (score: {content_score:.3f}) - page loaded successfully")
    else:
        print(f"  ⚠ Content check failed (score: {content_score:.3f}), but continuing...")
    
    # Input "sneakers"
    print("  ⌨️ Inputting 'sneakers'...")
    input_text_stealth("sneakers", APP_PACKAGE)
    time.sleep(1)
    
    # Step delay between Step 22 and Step 23
    step_delay(1, 2)
    
    # Step 23: Click region (0.128, 0.955), (0.206, 0.956), (0.208, 0.983), (0.128, 0.985)
    print("\n📍 [Step 23] Clicking region...")
    step23_region = {
        'x_min': min(0.128, 0.206, 0.208, 0.128),  # 0.128
        'x_max': max(0.128, 0.206, 0.208, 0.128),  # 0.208
        'y_min': min(0.955, 0.956, 0.983, 0.985),  # 0.955
        'y_max': max(0.955, 0.956, 0.983, 0.985)   # 0.985
    }
    click_random_in_rect(step23_region, "Step 23 Region")
    time.sleep(1)
    
    # Step delay between Step 23 and Step 24
    step_delay(1, 2)
    
    # Step 24: Check Excel E column (GenderEn) and click corresponding region
    print("\n📍 [Step 24] Checking GenderEn and clicking corresponding region...")
    gender = str(row_data.get('GenderEn', '')).lower().strip()
    print(f"  👤 Gender from Excel E column: '{gender}'")
    
    if gender == 'men':
        # Click men region: (0.028, 0.596), (0.397, 0.601), (0.403, 0.631), (0.042, 0.640)
        step24_region = {
            'x_min': min(0.028, 0.397, 0.403, 0.042),  # 0.028
            'x_max': max(0.028, 0.397, 0.403, 0.042),  # 0.403
            'y_min': min(0.596, 0.601, 0.631, 0.640),  # 0.596
            'y_max': max(0.596, 0.601, 0.631, 0.640)   # 0.640
        }
        print("  🎯 Clicking 'men' region...")
        click_random_in_rect(step24_region, "Step 24 Men Region")
    elif gender == 'women':
        # Click women region: (0.050, 0.684), (0.406, 0.691), (0.414, 0.715), (0.050, 0.721)
        step24_region = {
            'x_min': min(0.050, 0.406, 0.414, 0.050),  # 0.050
            'x_max': max(0.050, 0.406, 0.414, 0.050),  # 0.414
            'y_min': min(0.684, 0.691, 0.715, 0.721),  # 0.684
            'y_max': max(0.684, 0.691, 0.715, 0.721)   # 0.721
        }
        print("  🎯 Clicking 'women' region...")
        click_random_in_rect(step24_region, "Step 24 Women Region")
    else:
        print(f"  ⚠ Unknown gender '{gender}', defaulting to 'men' region...")
        # Default to men region
        step24_region = {
            'x_min': min(0.028, 0.397, 0.403, 0.042),  # 0.028
            'x_max': max(0.028, 0.397, 0.403, 0.042),  # 0.403
            'y_min': min(0.596, 0.601, 0.631, 0.640),  # 0.596
            'y_max': max(0.596, 0.601, 0.631, 0.640)   # 0.640
        }
        click_random_in_rect(step24_region, "Step 24 Default (Men) Region")
    
    time.sleep(1)
    
    # Step 24 to Step 25: Wait 7-9 seconds, then check for content changes
    print("\n📍 [Step 24-25] Waiting 7-9 seconds, then checking for content changes...")
    wait_time = random.uniform(7, 9)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before content check...")
    time.sleep(wait_time)
    
    # Content check region: (0.036, 0.244) to (0.061, 0.455)
    content_check_region_24_25 = {
        'x_min': min(0.036, 0.922, 0.911, 0.061),  # 0.036
        'x_max': max(0.036, 0.922, 0.911, 0.061),  # 0.922
        'y_min': min(0.244, 0.231, 0.444, 0.455),  # 0.231
        'y_max': max(0.244, 0.231, 0.444, 0.455)   # 0.455
    }
    
    # Retry click region: (0.808, 0.899) to (0.806, 0.934)
    retry_click_region_24_25 = {
        'x_min': min(0.808, 0.944, 0.950, 0.806),  # 0.806
        'x_max': max(0.808, 0.944, 0.950, 0.806),  # 0.950
        'y_min': min(0.899, 0.901, 0.927, 0.934),  # 0.899
        'y_max': max(0.899, 0.901, 0.927, 0.934)   # 0.934
    }
    
    # First attempt: Check if content is loaded (text, icons, color changes)
    print(f"  🔍 First check: Checking if region has content (text, icons, color changes)...")
    content_score, content_loaded = check_region_content_loaded(
        content_check_region_24_25,
        timeout=5,
        check_interval=0.5,
        threshold=0.3
    )
    
    if not content_loaded:
        # Retry: Click retry region and check again
        print(f"  ⚠ Content check failed (score: {content_score:.3f}), clicking retry region...")
        click_random_in_rect(retry_click_region_24_25, "Step 24-25 Retry Region")
        time.sleep(2)  # Wait after clicking
        
        # Second attempt: Check again
        print(f"  🔍 Second check: Checking content again...")
        content_score, content_loaded = check_region_content_loaded(
            content_check_region_24_25,
            timeout=5,
            check_interval=0.5,
            threshold=0.3
        )
        
        if not content_loaded:
            print(f"  ❌ Content check failed again (score: {content_score:.3f}), canceling task...")
            return False
    
    print(f"  ✓ Content check passed (score: {content_score:.3f}) - proceeding to Step 25")
    time.sleep(1)
    
    # Step delay between Step 24 and Step 25
    step_delay(1, 2)
    
    # Step 25: Click region and wait 3-4 seconds
    print("\n📍 [Step 25] Clicking region...")
    step25_region = {
        'x_min': min(0.842, 0.914, 0.914, 0.847),  # 0.842
        'x_max': max(0.842, 0.914, 0.914, 0.847),  # 0.914
        'y_min': min(0.471, 0.472, 0.489, 0.489),  # 0.471
        'y_max': max(0.471, 0.472, 0.489, 0.489)   # 0.489
    }
    click_random_in_rect(step25_region, "Step 25 Region")
    # Additional wait 3-4 seconds
    wait_time = random.uniform(3, 4)
    print(f"  ⏳ Additional waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step delay between Step 25 and Step 26
    step_delay(1, 2)
    
    # Step 26: Click region
    print("\n📍 [Step 26] Clicking region...")
    step26_region = {
        'x_min': min(0.083, 0.331, 0.333, 0.086),  # 0.083
        'x_max': max(0.083, 0.331, 0.333, 0.086),  # 0.333
        'y_min': min(0.331, 0.334, 0.364, 0.366),  # 0.331
        'y_max': max(0.331, 0.334, 0.364, 0.366)   # 0.366
    }
    click_random_in_rect(step26_region, "Step 26 Region")
    time.sleep(1)
    
    # Step delay between Step 26 and Step 27
    step_delay(1, 2)
    
    # Step 27: Click region
    print("\n📍 [Step 27] Clicking region...")
    step27_region = {
        'x_min': min(0.058, 0.381, 0.392, 0.056),  # 0.056
        'x_max': max(0.058, 0.381, 0.392, 0.056),  # 0.392
        'y_min': min(0.537, 0.544, 0.580, 0.589),  # 0.537
        'y_max': max(0.537, 0.544, 0.580, 0.589)   # 0.589
    }
    click_random_in_rect(step27_region, "Step 27 Region")
    time.sleep(1)
    
    # Step delay between Step 27 and Step 28
    step_delay(1, 2)
    
    # Step 28: Multi-step process for inputting "others"
    print("\n📍 [Step 28] Multi-step process for inputting 'others'...")
    
    # Step 28.1: Click first region (0.081, 0.412) to (0.253, 0.446)
    print("  🎯 [Step 28.1] Clicking first region...")
    step28_1_region = {
        'x_min': min(0.081, 0.253, 0.247, 0.081),  # 0.081
        'x_max': max(0.081, 0.253, 0.247, 0.081),  # 0.253
        'y_min': min(0.412, 0.417, 0.446, 0.448),  # 0.412
        'y_max': max(0.412, 0.417, 0.446, 0.448)   # 0.448
    }
    click_random_in_rect(step28_1_region, "Step 28.1 Region")
    
    # Wait 1-2 seconds
    wait_time = random.uniform(1, 2)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step 28.2: Click second region (0.161, 0.117) to (0.333, 0.158)
    print("  🎯 [Step 28.2] Clicking second region...")
    step28_2_region = {
        'x_min': min(0.161, 0.333, 0.322, 0.158),  # 0.158
        'x_max': max(0.161, 0.333, 0.322, 0.158),  # 0.333
        'y_min': min(0.117, 0.120, 0.149, 0.158),  # 0.117
        'y_max': max(0.117, 0.120, 0.149, 0.158)   # 0.158
    }
    click_random_in_rect(step28_2_region, "Step 28.2 Region")
    time.sleep(1)
    
    # Step 28.3: Input "others"
    print("  ⌨️ [Step 28.3] Inputting 'others'...")
    input_text_stealth("others", APP_PACKAGE)
    time.sleep(1)
    
    # Step 28.4: Click third region
    # Click region: 左上(47, 801) 右下(354, 912)
    print("  🎯 [Step 28.4] Clicking third region...")
    step28_4_region = pixels_to_rect(47, 801, 354, 912)
    click_random_in_rect(step28_4_region, "Step 28.4 Region")
    time.sleep(1)
    
    # Step delay between Step 28 and Step 29
    step_delay(1, 2)
    
    # Step 29: Click region
    # Click region: 左上(77, 1190) 右下(501, 1248)
    print("\n📍 [Step 29] Clicking region...")
    step29_region = pixels_to_rect(77, 1190, 501, 1248)
    click_random_in_rect(step29_region, "Step 29 Region")
    time.sleep(1)
    
    # Step 29.1: Input Excel I column (Brand)
    print("  ⌨️ [Step 29.1] Inputting Excel I column (Brand)...")
    brand = row_data.get('Brand', '')
    print(f"  ⌨️ Inputting Brand: {brand}")
    input_text_stealth(str(brand), APP_PACKAGE)
    time.sleep(1)
    
    # Step 29.1.1: Click region after Brand input
    # Click region: 左上(96, 2277) 右下(243, 2373)
    print("  🎯 [Step 29.1.1] Clicking region after Brand input...")
    step29_1_1_region = pixels_to_rect(96, 2277, 243, 2373)
    click_random_in_rect(step29_1_1_region, "Step 29.1.1 Region")
    time.sleep(1)
    
    # Step delay between Step 29 and Step 30
    step_delay(1, 2)
    
    # Step 30: Click input region
    print("\n📍 [Step 30] Clicking input region...")
    # Click region: 左上(60, 1411) 右下(617, 1500)
    step30_click_region = pixels_to_rect(60, 1411, 617, 1500)
    click_random_in_rect(step30_click_region, "Step 30 Input Region")
    
    # Wait 1-2 seconds
    wait_time = random.uniform(1, 2)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step delay between Step 30 and Step 33
    step_delay(1, 2)
    
    # Step 33: Click region and input size based on GenderEn
    print("\n📍 [Step 33] Clicking region and inputting size based on GenderEn...")
    
    # Step 33: Click input region first
    # Click region: 左上(146, 278) 右下(747, 369)
    print("  🎯 [Step 33] Clicking input region...")
    step33_click_region = pixels_to_rect(146, 278, 747, 369)
    click_random_in_rect(step33_click_region, "Step 33 Input Region")
    
    # Wait 1-2 seconds
    wait_time = random.uniform(1, 2)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before input...")
    time.sleep(wait_time)
    
    # Generate random size based on gender
    gender = str(row_data.get('GenderEn', '')).lower().strip()
    if gender == 'women':
        size = random.randint(36, 39)
        print(f"  👤 Gender: women, generating random size: {size}")
    elif gender == 'men':
        size = random.randint(40, 46)
        print(f"  👤 Gender: men, generating random size: {size}")
    else:
        # Default to men range if unknown
        size = random.randint(40, 46)
        print(f"  ⚠ Unknown gender '{gender}', defaulting to men size range: {size}")
    
    # Step 33: Input size
    print(f"  ⌨️ [Step 33] Inputting size: {size}")
    input_text_stealth(str(size), APP_PACKAGE)
    
    # Wait 2-3 seconds after input
    wait_time = random.uniform(2, 3)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds after input...")
    time.sleep(wait_time)
    
    # Step delay between Step 33 and Step 34
    step_delay(1, 2)
    
    # Step 34: Click region
    # Click region: 左上(18, 792) 右下(575, 902)
    print("\n📍 [Step 34] Clicking region...")
    step34_region = pixels_to_rect(18, 792, 575, 902)
    click_random_in_rect(step34_region, "Step 34 Region")
    time.sleep(1)
    
    # Step 34.5: Click additional region
    # Click region: 左上(317, 2184) 右下(713, 2287)
    print("\n📍 [Step 34.5] Clicking additional region...")
    step34_5_region = pixels_to_rect(317, 2184, 713, 2287)
    click_random_in_rect(step34_5_region, "Step 34.5 Region")
    time.sleep(1)
    
    # Step delay between Step 34 and Step 35
    step_delay(1, 2)
    
    # Step 35: Click region
    print("\n📍 [Step 35] Clicking region...")
    step35_region = {
        'x_min': min(0.839, 0.906, 0.908, 0.831),  # 0.831
        'x_max': max(0.839, 0.906, 0.908, 0.831),  # 0.908
        'y_min': min(0.635, 0.637, 0.660, 0.659),  # 0.635
        'y_max': max(0.635, 0.637, 0.660, 0.659)   # 0.660
    }
    click_random_in_rect(step35_region, "Step 35 Region")
    time.sleep(1)
    
    # Step delay between Step 35 and Step 36
    step_delay(1, 2)
    
    # Step 36: OCR detect any text and execute conditional logic
    print("\n📍 [Step 36] Checking for any text and executing conditional logic...")
    
    # OCR region for text detection: 左上(33, 566) 右下(627, 708)
    text_ocr_region = pixels_to_rect(33, 566, 627, 708)
    
    # Check for any text using OCR (low threshold for detection)
    text_found = False
    if initialize_ocr():
        screenshot = take_screenshot()
        if screenshot is not None:
            try:
                height, width = screenshot.shape[:2]
                x1 = int(text_ocr_region['x_min'] * width)
                x2 = int(text_ocr_region['x_max'] * width)
                y1 = int(text_ocr_region['y_min'] * height)
                y2 = int(text_ocr_region['y_max'] * height)
                region = screenshot[y1:y2, x1:x2]
                
                if region.size > 0:
                    gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                    enhanced = clahe.apply(gray_region)
                    # Low confidence threshold to detect any text
                    result = ocr_engine.readtext(enhanced, min_size=5)
                    
                    if result and len(result) > 0:
                        # If any text detected (low threshold)
                        for detection in result:
                            confidence = detection[2]
                            if confidence >= 0.1:  # Very low threshold
                                text_found = True
                                print(f"  ✓ Text found: '{detection[1]}' (confidence: {confidence:.3f})")
                                break
            except Exception as e:
                print(f"  ⚠ OCR error: {e}")
    
    if text_found:
        # Branch A: If any text found, click region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Branch A] Text found, clicking region...")
        step36_1_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_1_region, "Step 36.1 Region (Text Found)")
        time.sleep(1)
    else:
        # Branch B: If no text found, click three regions in sequence
        print("  🎯 [Branch B] No text found, clicking alternative regions...")
        
        # Step 36.2.1: Click first region 左上(23, 974) 右下(459, 1051)
        print("  🎯 [Step 36.2.1] Clicking first region...")
        step36_2_1_region = pixels_to_rect(23, 974, 459, 1051)
        click_random_in_rect(step36_2_1_region, "Step 36.2.1 Region")
        time.sleep(1)
        
        # Step 36.2.2: Click second region 左上(921, 504) 右下(1034, 559)
        print("  🎯 [Step 36.2.2] Clicking second region...")
        step36_2_2_region = pixels_to_rect(921, 504, 1034, 559)
        click_random_in_rect(step36_2_2_region, "Step 36.2.2 Region")
        time.sleep(1)
        
        # Step 36.2.3: Click third region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Step 36.2.3] Clicking third region...")
        step36_2_3_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_2_3_region, "Step 36.2.3 Region")
        time.sleep(1)
        
        # Step 36.2.4: Click fourth region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Step 36.2.4] Clicking fourth region...")
        step36_2_4_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_2_4_region, "Step 36.2.4 Region")
        time.sleep(1)
    
    # Step 36.5: Click final region and check for any text/pattern
    print("  🎯 [Step 36.5] Clicking final region...")
    step36_5_region = {
        'x_min': min(0.344, 0.664, 0.656, 0.350),  # 0.344
        'x_max': max(0.344, 0.664, 0.656, 0.350),  # 0.664
        'y_min': min(0.909, 0.914, 0.949, 0.946),  # 0.909
        'y_max': max(0.909, 0.914, 0.949, 0.946)   # 0.949
    }
    click_random_in_rect(step36_5_region, "Step 36.5 Region (Final)")
    
    # Wait 10-11 seconds
    wait_time = random.uniform(10, 11)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check...")
    time.sleep(wait_time)
    
    # OCR region for any text/pattern: 左上(141, 1802) 右下(1022, 2112)
    content_ocr_region = pixels_to_rect(141, 1802, 1022, 2112)
    
    def check_any_content(region_rect, confidence_threshold=0.1):
        """Check if any text or pattern exists in region (very low threshold)"""
        if not initialize_ocr():
            return False
        
        screenshot = take_screenshot()
        if screenshot is None:
            return False
        
        try:
            height, width = screenshot.shape[:2]
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                return False
            
            gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(gray_region)
            # Low confidence threshold to detect any text or pattern
            result = ocr_engine.readtext(enhanced, min_size=5)
            
            if result and len(result) > 0:
                for detection in result:
                    confidence = detection[2]
                    if confidence >= confidence_threshold:
                        print(f"  ✓ Content found: '{detection[1]}' (confidence: {confidence:.3f})")
                        return True
            
            return False
        except Exception as e:
            print(f"  ⚠ OCR error checking content: {e}")
            return False
    
    # First attempt: check for any content
    content_found = check_any_content(content_ocr_region, confidence_threshold=0.1)
    
    if not content_found:
        # Retry: click step36_5_region again and check
        print("  ⚠ No content found, retrying click and check...")
        click_random_in_rect(step36_5_region, "Step 36.5 Region (Retry)")
        wait_time = random.uniform(10, 11)
        print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check again...")
        time.sleep(wait_time)
        
        content_found = check_any_content(content_ocr_region, confidence_threshold=0.1)
        
        if not content_found:
            print("  ❌ No content found after retry, ending task...")
            return False
    
    # Step delay between Step 36 and Step 37
    step_delay(1, 2)
    
    # Helper function for Step 38 "Mark" detection
    def check_mark_text(region_rect, confidence_threshold=0.6):
        """Check if 'Mark' text exists in region with specified confidence threshold"""
        if not initialize_ocr():
            return False
        
        screenshot = take_screenshot()
        if screenshot is None:
            return False
        
        try:
            height, width = screenshot.shape[:2]
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                return False
            
            gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(gray_region)
            result = ocr_engine.readtext(enhanced)
            
            for detection in result:
                text = detection[1].lower()
                confidence = detection[2]
                if 'mark' in text and confidence >= confidence_threshold:
                    print(f"  ✓ 'Mark' text found: '{detection[1]}' (confidence: {confidence:.3f})")
                    return True
            
            return False
        except Exception as e:
            print(f"  ⚠ OCR error checking 'Mark': {e}")
            return False
    
    # Step 37: Click region (Final step)
    print("\n📍 [Step 37] Clicking region (Final step)...")
    step37_region = {
        'x_min': min(0.333, 0.697, 0.686, 0.331),  # 0.331
        'x_max': max(0.333, 0.697, 0.686, 0.331),  # 0.697
        'y_min': min(0.552, 0.561, 0.586, 0.581),  # 0.552
        'y_max': max(0.552, 0.561, 0.586, 0.581)   # 0.586
    }
    click_random_in_rect(step37_region, "Step 37 Region")
    
    # Wait 4-6 seconds after final click
    final_wait_time = random.uniform(4, 6)
    print(f"  ⏳ Waiting {final_wait_time:.1f} seconds after final click...")
    time.sleep(final_wait_time)
    print("  ✓ Final wait complete!")
    
    print("\n" + "="*60)
    print("✓ PHASE 0: IMAGE UPLOAD COMPLETE")
    print("="*60)
    
    return True


def phase1_create_listing(row_data):
    """
    Phase 1: Create initial listing with basic info
    Returns True if successful
    """
    print("\n=== PHASE 1: Creating Initial Listing ===")
    
    # 1. Click Sell button
    if not click_image("01_sell.png"):
        return False
    
    # 2. Handle permissions
    handle_permissions()
    
    # 3. Click category search
    if not click_image("04_category_search.png"):
        return False
    
    # 4. Click input field and type "others"
    if not click_image("05_input_others.png"):
        return False
    input_text_stealth("others", APP_PACKAGE)
    
    # 5. Select service others
    if not click_image("06_service_others.png"):
        return False
    
    # 6. Input title (based on region)
    if not click_image("07_title_input.png"):
        return False
    title = get_title_for_region(row_data, CURRENT_REGION)
    print(f"  📝 Using title for region {CURRENT_REGION}: {title}")
    input_text_stealth(title)
    
    # 7. Click condition menu (Phase 1)
    if not click_image("08_phase1_condition_menu.png"):
        return False
    # Assume first option is "New" - press ENTER
    d.shell("input keyevent KEYCODE_DPAD_DOWN")
    time.sleep(0.3)
    d.shell("input keyevent KEYCODE_ENTER")
    time.sleep(0.5)
    
    # 8. Input description (use same title)
    if not click_image("09_desc_input.png"):
        return False
    desc = title  # Using same title as description
    input_text_stealth(desc)
    
    # 9. Input price (based on region)
    if not click_image("10_price_input.png"):
        return False
    price = get_price_for_region(row_data, CURRENT_REGION)
    print(f"  💰 Using price for region {CURRENT_REGION}: {price}")
    input_text_stealth(str(price))
    
    # 10. Click delivery menu
    if not click_image("11_delivery_menu.png"):
        return False
    
    # 11. Submit
    if not click_image("12_submit.png"):
        return False
    
    time.sleep(3)  # Wait for listing to be created
    print("Phase 1 completed successfully")
    return True


def phase2_edit_details(row_data):
    """
    Phase 2: Edit listing to change category to sneakers and add details
    Returns True if successful
    """
    print("\n=== PHASE 2: Editing Details ===")
    
    # 12. Enter edit mode
    if not click_image("13_modify_page.png"):
        return False
    
    if not click_image("14_tab_inactive.png"):
        return False
    
    if not click_image("15_click_item.png"):
        return False
    
    if not click_image("16_menu_topright.png"):
        return False
    
    if not click_image("17_edit_listing_btn.png"):
        return False
    
    # 13. Change category to sneakers
    if not click_image("18_category_edit_input.png"):
        return False
    
    if not click_image("19_input_sneaker.png"):
        return False
    input_text_stealth("sneaker")
    
    # Handle gender-based category selection
    gender = str(row_data.get('GenderEn', '')).lower()
    print(f"Gender: {gender}")
    
    if gender == "men":
        # Look for men's sneaker category
        if click_image("20_cat_sneaker_men.png", timeout=5, required=False):
            pass  # Already clicked
        else:
            # Use keyboard navigation as fallback
            d.shell("input keyevent KEYCODE_DPAD_DOWN")
            time.sleep(0.3)
            d.shell("input keyevent KEYCODE_DPAD_DOWN")
            time.sleep(0.3)
            d.shell("input keyevent KEYCODE_ENTER")
            time.sleep(0.5)
    
    elif gender == "women":
        # Look for women's sneaker category
        if click_image("21_cat_sneaker_women.png", timeout=5, required=False):
            pass  # Already clicked
        else:
            # Use keyboard navigation as fallback
            d.shell("input keyevent KEYCODE_DPAD_DOWN")
            time.sleep(0.3)
            d.shell("input keyevent KEYCODE_ENTER")
            time.sleep(0.5)
    else:
        print(f"Unknown gender: {gender}, using default selection")
        d.shell("input keyevent KEYCODE_DPAD_DOWN")
        time.sleep(0.3)
        d.shell("input keyevent KEYCODE_ENTER")
        time.sleep(0.5)
    
    time.sleep(2)
    
    # 14. Edit attributes - Click the TOP edit button
    if not click_image("22_edit_details_top.png"):
        return False
    
    # Condition (Phase 2) - Brand New
    if not click_image("23_phase2_condition_menu.png"):
        return False
    
    if not click_image("24_phase2_option_brandnew.png"):
        return False
    
    # Brand Generic (Others)
    if not click_image("25_brand_menu.png"):
        return False
    
    if not click_image("26_brand_search_icon.png"):
        return False
    input_text_stealth("others", APP_PACKAGE)
    
    if not click_image("27_brand_select_others.png"):
        return False
    
    # Brand Specific
    if not click_image("28_brand_input.png"):
        return False
    brand = row_data.get('Brand', '')
    input_text_stealth(brand)
    
    # Scroll down
    scroll_down()
    
    # Size input
    if not click_image("30_size_input.png"):
        return False
    input_text_stealth("42")  # Default size, can be parameterized
    
    # Select first size option
    if not click_image("31_size_select_first.png"):
        return False
    
    # Save attributes
    if not click_image("32_save_btn.png"):
        return False
    
    time.sleep(2)
    print("Phase 2 completed successfully")
    return True


def phase3_delivery_activation(row_data):
    """
    Phase 3: Edit delivery options and activate listing
    Returns True if successful
    """
    print("\n=== PHASE 3: Delivery & Activation ===")
    
    # 15. Edit delivery (BOTTOM edit button)
    if not click_image("33_edit_delivery_bottom.png"):
        return False
    
    # Select meetup option
    if not click_image("34_delivery_meetup_option.png"):
        return False
    
    # Toggle meetup
    if not click_image("35_meetup_toggle.png"):
        return False
    
    # Location input
    if not click_image("36_location_input.png", timeout=5, required=False):
        print("Location input not found, continuing...")
    
    # Save delivery - try both save buttons
    if not click_image("37_save_delivery_1.png", timeout=3, required=False):
        click_image("38_save_delivery_2.png", timeout=3, required=False)
    
    time.sleep(1)
    
    # Final save (if needed)
    click_image("32_save_btn.png", timeout=3, required=False)
    
    # 16. Activate listing
    if not click_image("39_activate_btn.png"):
        return False
    
    if not click_image("40_confirm_activate.png"):
        return False
    
    time.sleep(3)  # Wait for activation
    print("Phase 3 completed successfully")
    return True


def call_vm_script(action, name, app_type=None, region=None, node=None):
    """
    调用 vm.sh 脚本执行虚拟机管理操作
    
    Args:
        action: 'load', 'save', 或 'new'
        name: 账号名称（BrowserID）
        app_type: 应用类型（仅 new 操作需要，默认 'Carousell'）
        region: 地区代码（仅 new 操作需要，默认 'GB'）
        node: 代理节点名称（仅 new 操作需要，默认 'UK-01'）
    
    Returns:
        True if successful, False otherwise
    """
    # vm.sh 脚本在设备上的路径
    # 使用相对路径，相对于设备当前工作目录（通常是 /data/local/tmp/）
    VM_SCRIPT_PATH = "./vm.sh"
    
    if not name or not name.strip():
        print(f"❌ [VM] Invalid name for {action} operation")
        return False
    
    try:
        # 转义特殊字符，防止命令注入
        name_escaped = shlex.quote(str(name).strip())
        
        if action == "load":
            # load 命令: ./vm.sh load <NAME>
            cmd = f"su -c '{VM_SCRIPT_PATH} load {name_escaped}'"
            print(f"\n🔄 [VM] Loading account: {name}")
        elif action == "save":
            # save 命令: ./vm.sh save <NAME>
            cmd = f"su -c '{VM_SCRIPT_PATH} save {name_escaped}'"
            print(f"\n💾 [VM] Saving account: {name}")
        elif action == "new":
            # new 命令: ./vm.sh new <NAME> [APP_TYPE] [REGION] [NODE]
            app_type = app_type or "Carousell"
            region = region or "GB"
            node = node or "UK-01"
            app_type_escaped = shlex.quote(str(app_type))
            region_escaped = shlex.quote(str(region))
            node_escaped = shlex.quote(str(node))
            cmd = f"su -c '{VM_SCRIPT_PATH} new {name_escaped} {app_type_escaped} {region_escaped} {node_escaped}'"
            print(f"\n🆕 [VM] Creating new account: {name} ({app_type}, {region}, {node})")
        else:
            print(f"❌ [VM] Unknown action: {action}")
            return False
        
        # 通过 adb shell 执行命令
        print(f"  📱 Executing: {cmd}")
        result = d.shell(cmd, timeout=120)  # 增加超时时间到 120 秒
        
        # 检查执行结果
        if result:
            result_str = str(result).strip()
            # 检查是否有错误信息
            if "错误" in result_str or "❌" in result_str or "Error" in result_str.lower():
                print(f"❌ [VM] {action} failed: {result_str}")
                return False
            else:
                print(f"✓ [VM] {action} completed successfully")
                # 等待操作完成（load 和 save 可能需要一些时间）
                if action == "load":
                    wait_time = 10  # load 需要更多时间等待应用启动
                    print(f"  ⏳ Waiting {wait_time} seconds for VM to load and app to start...")
                    time.sleep(wait_time)
                elif action == "save":
                    wait_time = 5  # save 通常较快
                    print(f"  ⏳ Waiting {wait_time} seconds for VM to save...")
                    time.sleep(wait_time)
                return True
        else:
            print(f"⚠ [VM] {action} returned no output, assuming success")
            # 即使没有输出也等待一段时间
            if action == "load":
                time.sleep(10)
            elif action == "save":
                time.sleep(5)
            return True
            
    except Exception as e:
        print(f"❌ [VM] Error executing {action}: {e}")
        import traceback
        traceback.print_exc()
        return False


def update_excel_status(excel_path, row_index, status_message):
    """Update Excel with status message in column A"""
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Row index in Excel (1-based, +2 for header and 0-based index)
        excel_row = row_index + 2
        
        # Write to column A
        ws.cell(row=excel_row, column=1, value=status_message)
        
        wb.save(excel_path)
        print(f"Updated Excel row {excel_row}: {status_message}")
    except Exception as e:
        print(f"Error updating Excel: {e}")


def process_listing(row_index, row_data):
    """
    Process a single listing through all phases
    Returns True if successful
    """
    sku = row_data.get('SKU', f'Row_{row_index}')
    browser_id = row_data.get('BrowserID', '')
    
    print(f"\n{'='*60}")
    print(f"Processing Row {row_index + 1}: SKU={sku}, BrowserID={browser_id}")
    print(f"{'='*60}")
    
    # Step 1: Load VM account (调用 vm.sh load)
    if not browser_id:
        print("⚠ Warning: BrowserID is empty, skipping VM load/save")
    else:
        if not call_vm_script("load", browser_id):
            error_msg = f"Error: Failed to load VM account {browser_id} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            update_excel_status(EXCEL_PATH, row_index, error_msg)
            print(f"❌ Failed to load VM account: {browser_id}")
            return False
    
    try:
        # Phase 0: Upload images (NEW)
        if not upload_images_phase(row_data):
            raise Exception("Phase 0 (Image Upload) failed")
        
        # Phase 1: Create initial listing
        if not phase1_create_listing(row_data):
            raise Exception("Phase 1 failed")
        
        # Phase 2: Edit details
        if not phase2_edit_details(row_data):
            raise Exception("Phase 2 failed")
        
        # Phase 3: Delivery and activation
        if not phase3_delivery_activation(row_data):
            raise Exception("Phase 3 failed")
        
        # Success
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status = f"Success: {timestamp}"
        update_excel_status(EXCEL_PATH, row_index, status)
        print(f"\n✓ Successfully processed SKU: {sku}")
        
        # Step 2: Save VM account (调用 vm.sh save)
        if browser_id:
            if not call_vm_script("save", browser_id):
                print(f"⚠ Warning: Failed to save VM account: {browser_id}, but listing was successful")
        
        return True
        
    except Exception as e:
        # Error
        error_msg = f"Error: {str(e)} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        update_excel_status(EXCEL_PATH, row_index, error_msg)
        print(f"\n✗ Failed to process SKU: {sku} - {e}")
        
        # 即使出错也尝试保存 VM 账号
        if browser_id:
            print(f"⚠ Attempting to save VM account after error: {browser_id}")
            call_vm_script("save", browser_id)
        
        return False


def main():
    """Main function to process all listings"""
    print("\n" + "="*60)
    print("CAROUSELL AUTO LISTING SCRIPT")
    print("="*60 + "\n")
    
    # Step 0: Using native Gboard with Smart Hybrid Input Strategy
    # No need to switch keyboard - using native Gboard for all input
    
    # Step 1: Detect current region
    global CURRENT_REGION
    CURRENT_REGION = detect_region()
    if not CURRENT_REGION:
        print("⚠ Warning: Could not detect region, defaulting to HK")
        CURRENT_REGION = 'HK'
    
    print(f"🌍 Current Region: {CURRENT_REGION}")
    print(f"   - Price Column: {'HKPrice' if CURRENT_REGION == 'HK' else 'MYPrice' if CURRENT_REGION == 'MY' else 'SGPrice'}")
    print(f"   - Title Column: {'ProductNameCn' if CURRENT_REGION == 'HK' else 'ProductNameEn'}")
    print()
    
    # Load Excel file
    print(f"Loading Excel file: {EXCEL_PATH}")
    df = pd.read_excel(EXCEL_PATH)
    print(f"Found {len(df)} rows to process\n")
    
    # Column mapping (0-based index)
    # A=0 (SKU/Status), B=1 (BrowserID), C=2 (ProductNameCn), D=3 (ProductNameEn), E=4 (GenderEn), 
    # F=5 (HKPrice), G=6 (MYPrice), H=7 (SGPrice), I=8 (Brand), J=9 (ImageFolder)
    # N=13 (Description content)
    df.columns = df.columns.str.strip()
    
    # Process each row
    success_count = 0
    error_count = 0
    
    for index, row in df.iterrows():
        # Convert row to dict for easier access
        row_data = {
            'SKU': row.iloc[0] if len(row) > 0 else '',
            'BrowserID': row.iloc[1] if len(row) > 1 else '',  # Column B (BrowserID)
            'ProductNameCn': row.iloc[2] if len(row) > 2 else '',  # Column C
            'ProductNameEn': row.iloc[3] if len(row) > 3 else '',  # Column D
            'GenderEn': row.iloc[4] if len(row) > 4 else '',      # Column E
            'HKPrice': row.iloc[5] if len(row) > 5 else '',       # Column F
            'MYPrice': row.iloc[6] if len(row) > 6 else '',        # Column G
            'SGPrice': row.iloc[7] if len(row) > 7 else '',        # Column H
            'Brand': row.iloc[8] if len(row) > 8 else '',          # Column I
            'ImageFolder': row.iloc[9] if len(row) > 9 else '',    # Column J
            'Description': row.iloc[13] if len(row) > 13 else ''  # Column N (index 13)
        }
        
        # Process the listing
        if process_listing(index, row_data):
            success_count += 1
        else:
            error_count += 1
        
        # Wait between listings
        print("\nWaiting 3 seconds before next listing...")
        time.sleep(3)
    
    # Summary
    print("\n" + "="*60)
    print("PROCESSING COMPLETE")
    print("="*60)
    print(f"Total Processed: {len(df)}")
    print(f"Successful: {success_count}")
    print(f"Errors: {error_count}")
    print("="*60 + "\n")


def test_from_step21(row_data):
    """
    Test function: Execute from Step 21 onwards to the end
    This function starts from Step 21 and continues through all remaining steps
    """
    print("\n" + "="*60)
    print("TEST: Starting from Step 21 to End")
    print("="*60 + "\n")
    
    # Step delay between Step 20 and Step 21
    step_delay(1, 2)
    
    # Step 21: Click region (0.072, 0.395), (0.622, 0.398), (0.614, 0.421), (0.083, 0.422)
    print("\n📍 [Step 21] Clicking region...")
    step21_region = {
        'x_min': min(0.072, 0.622, 0.614, 0.083),  # 0.072
        'x_max': max(0.072, 0.622, 0.614, 0.083),  # 0.622
        'y_min': min(0.395, 0.398, 0.421, 0.422),  # 0.395
        'y_max': max(0.395, 0.398, 0.421, 0.422)   # 0.422
    }
    click_random_in_rect(step21_region, "Step 21 Region")
    time.sleep(1)
    
    # Step delay between Step 21 and Step 22
    step_delay(1, 2)
    
    # Step 22: Use Step 9.1 content check logic, then input "sneakers"
    print("\n📍 [Step 22] Checking content (using Step 9.1 logic), then inputting 'sneakers'...")
    
    # Use the same content check region as Step 9.1: (0.039, 0.412) to (0.992, 0.969)
    content_check_region_step22 = {
        'x_min': 0.039,
        'x_max': 0.992,
        'y_min': 0.412,
        'y_max': 0.969
    }
    
    # Wait 6-9 seconds before checking (same as Step 9.1)
    wait_time = random.uniform(6, 9)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before content check...")
    time.sleep(wait_time)
    
    # Check if content is loaded (using same logic as Step 9.1)
    print(f"  🔍 Starting content check (to verify network status and page load)...")
    content_score, content_loaded = check_region_content_loaded(
        content_check_region_step22,
        timeout=10,
        check_interval=0.5
    )
    
    if content_loaded:
        print(f"  ✓ Content check passed (score: {content_score:.3f}) - page loaded successfully")
    else:
        print(f"  ⚠ Content check failed (score: {content_score:.3f}), but continuing...")
    
    # Input "sneakers"
    print("  ⌨️ Inputting 'sneakers'...")
    input_text_stealth("sneakers", APP_PACKAGE)
    time.sleep(1)
    
    # Step delay between Step 22 and Step 23
    step_delay(1, 2)
    
    # Step 23: Click region (0.128, 0.955), (0.206, 0.956), (0.208, 0.983), (0.128, 0.985)
    print("\n📍 [Step 23] Clicking region...")
    step23_region = {
        'x_min': min(0.128, 0.206, 0.208, 0.128),  # 0.128
        'x_max': max(0.128, 0.206, 0.208, 0.128),  # 0.208
        'y_min': min(0.955, 0.956, 0.983, 0.985),  # 0.955
        'y_max': max(0.955, 0.956, 0.983, 0.985)   # 0.985
    }
    click_random_in_rect(step23_region, "Step 23 Region")
    time.sleep(1)
    
    # Step delay between Step 23 and Step 24
    step_delay(1, 2)
    
    # Step 24: Check Excel E column (GenderEn) and click corresponding region
    print("\n📍 [Step 24] Checking GenderEn and clicking corresponding region...")
    gender = str(row_data.get('GenderEn', '')).lower().strip()
    print(f"  👤 Gender from Excel E column: '{gender}'")
    
    if gender == 'men':
        # Click men region: (0.028, 0.596), (0.397, 0.601), (0.403, 0.631), (0.042, 0.640)
        step24_region = {
            'x_min': min(0.028, 0.397, 0.403, 0.042),  # 0.028
            'x_max': max(0.028, 0.397, 0.403, 0.042),  # 0.403
            'y_min': min(0.596, 0.601, 0.631, 0.640),  # 0.596
            'y_max': max(0.596, 0.601, 0.631, 0.640)   # 0.640
        }
        print("  🎯 Clicking 'men' region...")
        click_random_in_rect(step24_region, "Step 24 Men Region")
    elif gender == 'women':
        # Click women region: (0.050, 0.684), (0.406, 0.691), (0.414, 0.715), (0.050, 0.721)
        step24_region = {
            'x_min': min(0.050, 0.406, 0.414, 0.050),  # 0.050
            'x_max': max(0.050, 0.406, 0.414, 0.050),  # 0.414
            'y_min': min(0.684, 0.691, 0.715, 0.721),  # 0.684
            'y_max': max(0.684, 0.691, 0.715, 0.721)   # 0.721
        }
        print("  🎯 Clicking 'women' region...")
        click_random_in_rect(step24_region, "Step 24 Women Region")
    else:
        print(f"  ⚠ Unknown gender '{gender}', defaulting to 'men' region...")
        # Default to men region
        step24_region = {
            'x_min': min(0.028, 0.397, 0.403, 0.042),  # 0.028
            'x_max': max(0.028, 0.397, 0.403, 0.042),  # 0.403
            'y_min': min(0.596, 0.601, 0.631, 0.640),  # 0.596
            'y_max': max(0.596, 0.601, 0.631, 0.640)   # 0.640
        }
        click_random_in_rect(step24_region, "Step 24 Default (Men) Region")
    
    time.sleep(1)
    
    # Step 24 to Step 25: Wait 7-9 seconds, then check for content changes
    print("\n📍 [Step 24-25] Waiting 7-9 seconds, then checking for content changes...")
    wait_time = random.uniform(7, 9)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before content check...")
    time.sleep(wait_time)
    
    # Content check region: (0.036, 0.244) to (0.061, 0.455)
    content_check_region_24_25 = {
        'x_min': min(0.036, 0.922, 0.911, 0.061),  # 0.036
        'x_max': max(0.036, 0.922, 0.911, 0.061),  # 0.922
        'y_min': min(0.244, 0.231, 0.444, 0.455),  # 0.231
        'y_max': max(0.244, 0.231, 0.444, 0.455)   # 0.455
    }
    
    # Retry click region: (0.808, 0.899) to (0.806, 0.934)
    retry_click_region_24_25 = {
        'x_min': min(0.808, 0.944, 0.950, 0.806),  # 0.806
        'x_max': max(0.808, 0.944, 0.950, 0.806),  # 0.950
        'y_min': min(0.899, 0.901, 0.927, 0.934),  # 0.899
        'y_max': max(0.899, 0.901, 0.927, 0.934)   # 0.934
    }
    
    # First attempt: Check if content is loaded (text, icons, color changes)
    print(f"  🔍 First check: Checking if region has content (text, icons, color changes)...")
    content_score, content_loaded = check_region_content_loaded(
        content_check_region_24_25,
        timeout=5,
        check_interval=0.5,
        threshold=0.3
    )
    
    if not content_loaded:
        # Retry: Click retry region and check again
        print(f"  ⚠ Content check failed (score: {content_score:.3f}), clicking retry region...")
        click_random_in_rect(retry_click_region_24_25, "Step 24-25 Retry Region")
        time.sleep(2)  # Wait after clicking
        
        # Second attempt: Check again
        print(f"  🔍 Second check: Checking content again...")
        content_score, content_loaded = check_region_content_loaded(
            content_check_region_24_25,
            timeout=5,
            check_interval=0.5,
            threshold=0.3
        )
        
        if not content_loaded:
            print(f"  ❌ Content check failed again (score: {content_score:.3f}), canceling task...")
            return False
    
    print(f"  ✓ Content check passed (score: {content_score:.3f}) - proceeding to Step 25")
    time.sleep(1)
    
    # Step delay between Step 24 and Step 25
    step_delay(1, 2)
    
    # Step 25: Click region and wait 3-4 seconds
    print("\n📍 [Step 25] Clicking region...")
    step25_region = {
        'x_min': min(0.842, 0.914, 0.914, 0.847),  # 0.842
        'x_max': max(0.842, 0.914, 0.914, 0.847),  # 0.914
        'y_min': min(0.471, 0.472, 0.489, 0.489),  # 0.471
        'y_max': max(0.471, 0.472, 0.489, 0.489)   # 0.489
    }
    click_random_in_rect(step25_region, "Step 25 Region")
    # Additional wait 3-4 seconds
    wait_time = random.uniform(3, 4)
    print(f"  ⏳ Additional waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step delay between Step 25 and Step 26
    step_delay(1, 2)
    
    # Step 26: Click region
    print("\n📍 [Step 26] Clicking region...")
    step26_region = {
        'x_min': min(0.083, 0.331, 0.333, 0.086),  # 0.083
        'x_max': max(0.083, 0.331, 0.333, 0.086),  # 0.333
        'y_min': min(0.331, 0.334, 0.364, 0.366),  # 0.331
        'y_max': max(0.331, 0.334, 0.364, 0.366)   # 0.366
    }
    click_random_in_rect(step26_region, "Step 26 Region")
    time.sleep(1)
    
    # Step delay between Step 26 and Step 27
    step_delay(1, 2)
    
    # Step 27: Click region
    print("\n📍 [Step 27] Clicking region...")
    step27_region = {
        'x_min': min(0.058, 0.381, 0.392, 0.056),  # 0.056
        'x_max': max(0.058, 0.381, 0.392, 0.056),  # 0.392
        'y_min': min(0.537, 0.544, 0.580, 0.589),  # 0.537
        'y_max': max(0.537, 0.544, 0.580, 0.589)   # 0.589
    }
    click_random_in_rect(step27_region, "Step 27 Region")
    time.sleep(1)
    
    # Step delay between Step 27 and Step 28
    step_delay(1, 2)
    
    # Step 28: Multi-step process for inputting "others"
    print("\n📍 [Step 28] Multi-step process for inputting 'others'...")
    
    # Step 28.1: Click first region (0.081, 0.412) to (0.253, 0.446)
    print("  🎯 [Step 28.1] Clicking first region...")
    step28_1_region = {
        'x_min': min(0.081, 0.253, 0.247, 0.081),  # 0.081
        'x_max': max(0.081, 0.253, 0.247, 0.081),  # 0.253
        'y_min': min(0.412, 0.417, 0.446, 0.448),  # 0.412
        'y_max': max(0.412, 0.417, 0.446, 0.448)   # 0.448
    }
    click_random_in_rect(step28_1_region, "Step 28.1 Region")
    
    # Wait 1-2 seconds
    wait_time = random.uniform(1, 2)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step 28.2: Click second region (0.161, 0.117) to (0.333, 0.158)
    print("  🎯 [Step 28.2] Clicking second region...")
    step28_2_region = {
        'x_min': min(0.161, 0.333, 0.322, 0.158),  # 0.158
        'x_max': max(0.161, 0.333, 0.322, 0.158),  # 0.333
        'y_min': min(0.117, 0.120, 0.149, 0.158),  # 0.117
        'y_max': max(0.117, 0.120, 0.149, 0.158)   # 0.158
    }
    click_random_in_rect(step28_2_region, "Step 28.2 Region")
    time.sleep(1)
    
    # Step 28.3: Input "others"
    print("  ⌨️ [Step 28.3] Inputting 'others'...")
    input_text_stealth("others", APP_PACKAGE)
    time.sleep(1)
    
    # Step 28.4: Click third region
    # Click region: 左上(47, 801) 右下(354, 912)
    print("  🎯 [Step 28.4] Clicking third region...")
    step28_4_region = pixels_to_rect(47, 801, 354, 912)
    click_random_in_rect(step28_4_region, "Step 28.4 Region")
    time.sleep(1)
    
    # Step delay between Step 28 and Step 29
    step_delay(1, 2)
    
    # Step 29: Click region
    # Click region: 左上(77, 1190) 右下(501, 1248)
    print("\n📍 [Step 29] Clicking region...")
    step29_region = pixels_to_rect(77, 1190, 501, 1248)
    click_random_in_rect(step29_region, "Step 29 Region")
    time.sleep(1)
    
    # Step 29.1: Input Excel I column (Brand)
    print("  ⌨️ [Step 29.1] Inputting Excel I column (Brand)...")
    brand = row_data.get('Brand', '')
    print(f"  ⌨️ Inputting Brand: {brand}")
    input_text_stealth(str(brand), APP_PACKAGE)
    time.sleep(1)
    
    # Step 29.1.1: Click region after Brand input
    # Click region: 左上(96, 2277) 右下(243, 2373)
    print("  🎯 [Step 29.1.1] Clicking region after Brand input...")
    step29_1_1_region = pixels_to_rect(96, 2277, 243, 2373)
    click_random_in_rect(step29_1_1_region, "Step 29.1.1 Region")
    time.sleep(1)
    
    # Step delay between Step 29 and Step 30
    step_delay(1, 2)
    
    # Step 30: Click input region
    print("\n📍 [Step 30] Clicking input region...")
    # Click region: 左上(60, 1411) 右下(617, 1500)
    step30_click_region = pixels_to_rect(60, 1411, 617, 1500)
    click_random_in_rect(step30_click_region, "Step 30 Input Region")
    
    # Wait 1-2 seconds
    wait_time = random.uniform(1, 2)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step delay between Step 30 and Step 33
    step_delay(1, 2)
    
    # Step 33: Click region and input size based on GenderEn
    print("\n📍 [Step 33] Clicking region and inputting size based on GenderEn...")
    
    # Step 33: Click input region first
    # Click region: 左上(146, 278) 右下(747, 369)
    print("  🎯 [Step 33] Clicking input region...")
    step33_click_region = pixels_to_rect(146, 278, 747, 369)
    click_random_in_rect(step33_click_region, "Step 33 Input Region")
    
    # Wait 1-2 seconds
    wait_time = random.uniform(1, 2)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before input...")
    time.sleep(wait_time)
    
    # Generate random size based on gender
    gender = str(row_data.get('GenderEn', '')).lower().strip()
    if gender == 'women':
        size = random.randint(36, 39)
        print(f"  👤 Gender: women, generating random size: {size}")
    elif gender == 'men':
        size = random.randint(40, 46)
        print(f"  👤 Gender: men, generating random size: {size}")
    else:
        # Default to men range if unknown
        size = random.randint(40, 46)
        print(f"  ⚠ Unknown gender '{gender}', defaulting to men size range: {size}")
    
    # Step 33: Input size
    print(f"  ⌨️ [Step 33] Inputting size: {size}")
    input_text_stealth(str(size), APP_PACKAGE)
    
    # Wait 2-3 seconds after input
    wait_time = random.uniform(2, 3)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds after input...")
    time.sleep(wait_time)
    
    # Step delay between Step 33 and Step 34
    step_delay(1, 2)
    
    # Step 34: Click region
    # Click region: 左上(18, 792) 右下(575, 902)
    print("\n📍 [Step 34] Clicking region...")
    step34_region = pixels_to_rect(18, 792, 575, 902)
    click_random_in_rect(step34_region, "Step 34 Region")
    time.sleep(1)
    
    # Step 34.5: Click additional region
    # Click region: 左上(317, 2184) 右下(713, 2287)
    print("\n📍 [Step 34.5] Clicking additional region...")
    step34_5_region = pixels_to_rect(317, 2184, 713, 2287)
    click_random_in_rect(step34_5_region, "Step 34.5 Region")
    time.sleep(1)
    
    # Step delay between Step 34 and Step 35
    step_delay(1, 2)
    
    # Step 35: Click region
    print("\n📍 [Step 35] Clicking region...")
    step35_region = {
        'x_min': min(0.839, 0.906, 0.908, 0.831),  # 0.831
        'x_max': max(0.839, 0.906, 0.908, 0.831),  # 0.908
        'y_min': min(0.635, 0.637, 0.660, 0.659),  # 0.635
        'y_max': max(0.635, 0.637, 0.660, 0.659)   # 0.660
    }
    click_random_in_rect(step35_region, "Step 35 Region")
    time.sleep(1)
    
    # Step delay between Step 35 and Step 36
    step_delay(1, 2)
    
    # Step 36: OCR detect any text and execute conditional logic
    print("\n📍 [Step 36] Checking for any text and executing conditional logic...")
    
    # OCR region for text detection: 左上(33, 566) 右下(627, 708)
    text_ocr_region = pixels_to_rect(33, 566, 627, 708)
    
    # Check for any text using OCR (low threshold for detection)
    text_found = False
    if initialize_ocr():
        screenshot = take_screenshot()
        if screenshot is not None:
            try:
                height, width = screenshot.shape[:2]
                x1 = int(text_ocr_region['x_min'] * width)
                x2 = int(text_ocr_region['x_max'] * width)
                y1 = int(text_ocr_region['y_min'] * height)
                y2 = int(text_ocr_region['y_max'] * height)
                region = screenshot[y1:y2, x1:x2]
                
                if region.size > 0:
                    gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                    enhanced = clahe.apply(gray_region)
                    # Low confidence threshold to detect any text
                    result = ocr_engine.readtext(enhanced, min_size=5)
                    
                    if result and len(result) > 0:
                        # If any text detected (low threshold)
                        for detection in result:
                            confidence = detection[2]
                            if confidence >= 0.1:  # Very low threshold
                                text_found = True
                                print(f"  ✓ Text found: '{detection[1]}' (confidence: {confidence:.3f})")
                                break
            except Exception as e:
                print(f"  ⚠ OCR error: {e}")
    
    if text_found:
        # Branch A: If any text found, click region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Branch A] Text found, clicking region...")
        step36_1_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_1_region, "Step 36.1 Region (Text Found)")
        time.sleep(1)
    else:
        # Branch B: If no text found, click three regions in sequence
        print("  🎯 [Branch B] No text found, clicking alternative regions...")
        
        # Step 36.2.1: Click first region 左上(23, 974) 右下(459, 1051)
        print("  🎯 [Step 36.2.1] Clicking first region...")
        step36_2_1_region = pixels_to_rect(23, 974, 459, 1051)
        click_random_in_rect(step36_2_1_region, "Step 36.2.1 Region")
        time.sleep(1)
        
        # Step 36.2.2: Click second region 左上(921, 504) 右下(1034, 559)
        print("  🎯 [Step 36.2.2] Clicking second region...")
        step36_2_2_region = pixels_to_rect(921, 504, 1034, 559)
        click_random_in_rect(step36_2_2_region, "Step 36.2.2 Region")
        time.sleep(1)
        
        # Step 36.2.3: Click third region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Step 36.2.3] Clicking third region...")
        step36_2_3_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_2_3_region, "Step 36.2.3 Region")
        time.sleep(1)
        
        # Step 36.2.4: Click fourth region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Step 36.2.4] Clicking fourth region...")
        step36_2_4_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_2_4_region, "Step 36.2.4 Region")
        time.sleep(1)
    
    # Step 36.5: Click final region and check for any text/pattern
    print("  🎯 [Step 36.5] Clicking final region...")
    step36_5_region = {
        'x_min': min(0.344, 0.664, 0.656, 0.350),  # 0.344
        'x_max': max(0.344, 0.664, 0.656, 0.350),  # 0.664
        'y_min': min(0.909, 0.914, 0.949, 0.946),  # 0.909
        'y_max': max(0.909, 0.914, 0.949, 0.946)   # 0.949
    }
    click_random_in_rect(step36_5_region, "Step 36.5 Region (Final)")
    
    # Wait 10-11 seconds
    wait_time = random.uniform(10, 11)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check...")
    time.sleep(wait_time)
    
    # OCR region for any text/pattern: 左上(141, 1802) 右下(1022, 2112)
    content_ocr_region = pixels_to_rect(141, 1802, 1022, 2112)
    
    def check_any_content(region_rect, confidence_threshold=0.1):
        """Check if any text or pattern exists in region (very low threshold)"""
        if not initialize_ocr():
            return False
        
        screenshot = take_screenshot()
        if screenshot is None:
            return False
        
        try:
            height, width = screenshot.shape[:2]
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                return False
            
            gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(gray_region)
            # Low confidence threshold to detect any text or pattern
            result = ocr_engine.readtext(enhanced, min_size=5)
            
            if result and len(result) > 0:
                for detection in result:
                    confidence = detection[2]
                    if confidence >= confidence_threshold:
                        print(f"  ✓ Content found: '{detection[1]}' (confidence: {confidence:.3f})")
                        return True
            
            return False
        except Exception as e:
            print(f"  ⚠ OCR error checking content: {e}")
            return False
    
    # First attempt: check for any content
    content_found = check_any_content(content_ocr_region, confidence_threshold=0.1)
    
    if not content_found:
        # Retry: click step36_5_region again and check
        print("  ⚠ No content found, retrying click and check...")
        click_random_in_rect(step36_5_region, "Step 36.5 Region (Retry)")
        wait_time = random.uniform(10, 11)
        print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check again...")
        time.sleep(wait_time)
        
        content_found = check_any_content(content_ocr_region, confidence_threshold=0.1)
        
        if not content_found:
            print("  ❌ No content found after retry, ending task...")
            return False
    
    # Step delay between Step 36 and Step 37
    step_delay(1, 2)
    
    # Helper function for Step 38 "Mark" detection
    def check_mark_text(region_rect, confidence_threshold=0.6):
        """Check if 'Mark' text exists in region with specified confidence threshold"""
        if not initialize_ocr():
            return False
        
        screenshot = take_screenshot()
        if screenshot is None:
            return False
        
        try:
            height, width = screenshot.shape[:2]
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                return False
            
            gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(gray_region)
            result = ocr_engine.readtext(enhanced)
            
            for detection in result:
                text = detection[1].lower()
                confidence = detection[2]
                if 'mark' in text and confidence >= confidence_threshold:
                    print(f"  ✓ 'Mark' text found: '{detection[1]}' (confidence: {confidence:.3f})")
                    return True
            
            return False
        except Exception as e:
            print(f"  ⚠ OCR error checking 'Mark': {e}")
            return False
    
    # Step 37: Click region
    print("\n📍 [Step 37] Clicking region...")
    step37_region = {
        'x_min': min(0.333, 0.697, 0.686, 0.331),  # 0.331
        'x_max': max(0.333, 0.697, 0.686, 0.331),  # 0.697
        'y_min': min(0.552, 0.561, 0.586, 0.581),  # 0.552
        'y_max': max(0.552, 0.561, 0.586, 0.581)   # 0.586
    }
    click_random_in_rect(step37_region, "Step 37 Region")
    time.sleep(1)
    
    # Step delay between Step 37 and Step 38
    step_delay(1, 2)
    
    # Step 38: Click region, wait 5-6 seconds, then OCR for any content
    print("\n📍 [Step 38] Clicking region, waiting, then checking for any content...")
    step38_1_region = {
        'x_min': min(0.633, 0.808, 0.808, 0.622),  # 0.622
        'x_max': max(0.633, 0.808, 0.808, 0.622),  # 0.808
        'y_min': min(0.534, 0.536, 0.557, 0.560),  # 0.534
        'y_max': max(0.534, 0.536, 0.557, 0.560)   # 0.560
    }
    click_random_in_rect(step38_1_region, "Step 38.1 Region")
    
    # Wait 5-6 seconds
    wait_time = random.uniform(5, 6)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check...")
    time.sleep(wait_time)
    
    # OCR region for any content: 左上(354, 1809) 右下(1053, 2116)
    step38_content_ocr_region = pixels_to_rect(354, 1809, 1053, 2116)
    
    # Check for any text or pattern (very low threshold)
    if initialize_ocr():
        screenshot = take_screenshot()
        if screenshot is not None:
            try:
                height, width = screenshot.shape[:2]
                x1 = int(step38_content_ocr_region['x_min'] * width)
                x2 = int(step38_content_ocr_region['x_max'] * width)
                y1 = int(step38_content_ocr_region['y_min'] * height)
                y2 = int(step38_content_ocr_region['y_max'] * height)
                region = screenshot[y1:y2, x1:x2]
                
                if region.size > 0:
                    gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                    enhanced = clahe.apply(gray_region)
                    # Low confidence threshold to detect any text or pattern
                    result = ocr_engine.readtext(enhanced, min_size=5)
                    
                    content_found = False
                    if result and len(result) > 0:
                        for detection in result:
                            confidence = detection[2]
                            if confidence >= 0.1:  # Very low threshold
                                content_found = True
                                print(f"  ✓ Content found: '{detection[1]}' (confidence: {confidence:.3f})")
                                break
                    
                    if content_found:
                        print("  ✓ Content detected, proceeding to next cycle!")
                    else:
                        print("  ⚠ No content found, but continuing to next cycle...")
            except Exception as e:
                print(f"  ⚠ OCR error: {e}, but continuing to next cycle...")
    
    print("\n" + "="*60)
    print("✓ TEST COMPLETE: All steps from Step 21 to Step 38 completed")
    print("="*60 + "\n")
    
    return True


def test_from_step33(row_data):
    """
    Test function: Execute from Step 33 onwards to the end
    This function starts from Step 33 and continues through all remaining steps
    """
    print("\n" + "="*60)
    print("TEST: Starting from Step 33 to End")
    print("="*60 + "\n")
    
    # Step 33: Click region and input size based on GenderEn
    print("\n📍 [Step 33] Clicking region and inputting size based on GenderEn...")
    
    # Step 33: Click input region first
    # Click region: 左上(146, 278) 右下(747, 369)
    print("  🎯 [Step 33] Clicking input region...")
    step33_click_region = pixels_to_rect(146, 278, 747, 369)
    click_random_in_rect(step33_click_region, "Step 33 Input Region")
    
    # Wait 1-2 seconds
    wait_time = random.uniform(1, 2)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before input...")
    time.sleep(wait_time)
    
    # Generate random size based on gender
    gender = str(row_data.get('GenderEn', '')).lower().strip()
    if gender == 'women':
        size = random.randint(36, 39)
        print(f"  👤 Gender: women, generating random size: {size}")
    elif gender == 'men':
        size = random.randint(40, 46)
        print(f"  👤 Gender: men, generating random size: {size}")
    else:
        # Default to men range if unknown
        size = random.randint(40, 46)
        print(f"  ⚠ Unknown gender '{gender}', defaulting to men size range: {size}")
    
    # Step 33: Input size
    print(f"  ⌨️ [Step 33] Inputting size: {size}")
    input_text_stealth(str(size), APP_PACKAGE)
    
    # Wait 2-3 seconds after input
    wait_time = random.uniform(2, 3)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds after input...")
    time.sleep(wait_time)
    
    # Step delay between Step 33 and Step 34
    step_delay(1, 2)
    
    # Step 34: Click region
    # Click region: 左上(18, 792) 右下(575, 902)
    print("\n📍 [Step 34] Clicking region...")
    step34_region = pixels_to_rect(18, 792, 575, 902)
    click_random_in_rect(step34_region, "Step 34 Region")
    time.sleep(1)
    
    # Step 34.5: Click additional region
    # Click region: 左上(317, 2184) 右下(713, 2287)
    print("\n📍 [Step 34.5] Clicking additional region...")
    step34_5_region = pixels_to_rect(317, 2184, 713, 2287)
    click_random_in_rect(step34_5_region, "Step 34.5 Region")
    time.sleep(1)
    
    # Step delay between Step 34 and Step 35
    step_delay(1, 2)
    
    # Step 35: Click region
    print("\n📍 [Step 35] Clicking region...")
    step35_region = {
        'x_min': min(0.839, 0.906, 0.908, 0.831),  # 0.831
        'x_max': max(0.839, 0.906, 0.908, 0.831),  # 0.908
        'y_min': min(0.635, 0.637, 0.660, 0.659),  # 0.635
        'y_max': max(0.635, 0.637, 0.660, 0.659)   # 0.660
    }
    click_random_in_rect(step35_region, "Step 35 Region")
    time.sleep(1)
    
    # Step delay between Step 35 and Step 36
    step_delay(1, 2)
    
    # Step 36: OCR detect any text and execute conditional logic
    print("\n📍 [Step 36] Checking for any text and executing conditional logic...")
    
    # OCR region for text detection: 左上(33, 566) 右下(627, 708)
    text_ocr_region = pixels_to_rect(33, 566, 627, 708)
    
    # Check for any text using OCR (low threshold for detection)
    text_found = False
    if initialize_ocr():
        screenshot = take_screenshot()
        if screenshot is not None:
            try:
                height, width = screenshot.shape[:2]
                x1 = int(text_ocr_region['x_min'] * width)
                x2 = int(text_ocr_region['x_max'] * width)
                y1 = int(text_ocr_region['y_min'] * height)
                y2 = int(text_ocr_region['y_max'] * height)
                region = screenshot[y1:y2, x1:x2]
                
                if region.size > 0:
                    gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                    enhanced = clahe.apply(gray_region)
                    # Low confidence threshold to detect any text
                    result = ocr_engine.readtext(enhanced, min_size=5)
                    
                    if result and len(result) > 0:
                        # If any text detected (low threshold)
                        for detection in result:
                            confidence = detection[2]
                            if confidence >= 0.1:  # Very low threshold
                                text_found = True
                                print(f"  ✓ Text found: '{detection[1]}' (confidence: {confidence:.3f})")
                                break
            except Exception as e:
                print(f"  ⚠ OCR error: {e}")
    
    if text_found:
        # Branch A: If any text found, click region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Branch A] Text found, clicking region...")
        step36_1_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_1_region, "Step 36.1 Region (Text Found)")
        time.sleep(1)
    else:
        # Branch B: If no text found, click three regions in sequence
        print("  🎯 [Branch B] No text found, clicking alternative regions...")
        
        # Step 36.2.1: Click first region 左上(23, 974) 右下(459, 1051)
        print("  🎯 [Step 36.2.1] Clicking first region...")
        step36_2_1_region = pixels_to_rect(23, 974, 459, 1051)
        click_random_in_rect(step36_2_1_region, "Step 36.2.1 Region")
        time.sleep(1)
        
        # Step 36.2.2: Click second region 左上(921, 504) 右下(1034, 559)
        print("  🎯 [Step 36.2.2] Clicking second region...")
        step36_2_2_region = pixels_to_rect(921, 504, 1034, 559)
        click_random_in_rect(step36_2_2_region, "Step 36.2.2 Region")
        time.sleep(1)
        
        # Step 36.2.3: Click third region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Step 36.2.3] Clicking third region...")
        step36_2_3_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_2_3_region, "Step 36.2.3 Region")
        time.sleep(1)
        
        # Step 36.2.4: Click fourth region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Step 36.2.4] Clicking fourth region...")
        step36_2_4_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_2_4_region, "Step 36.2.4 Region")
        time.sleep(1)
    
    # Step 36.5: Click final region and check for any text/pattern
    print("  🎯 [Step 36.5] Clicking final region...")
    step36_5_region = {
        'x_min': min(0.344, 0.664, 0.656, 0.350),  # 0.344
        'x_max': max(0.344, 0.664, 0.656, 0.350),  # 0.664
        'y_min': min(0.909, 0.914, 0.949, 0.946),  # 0.909
        'y_max': max(0.909, 0.914, 0.949, 0.946)   # 0.949
    }
    click_random_in_rect(step36_5_region, "Step 36.5 Region (Final)")
    
    # Wait 10-11 seconds
    wait_time = random.uniform(10, 11)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check...")
    time.sleep(wait_time)
    
    # OCR region for any text/pattern: 左上(141, 1802) 右下(1022, 2112)
    content_ocr_region = pixels_to_rect(141, 1802, 1022, 2112)
    
    def check_any_content(region_rect, confidence_threshold=0.1):
        """Check if any text or pattern exists in region (very low threshold)"""
        if not initialize_ocr():
            return False
        
        screenshot = take_screenshot()
        if screenshot is None:
            return False
        
        try:
            height, width = screenshot.shape[:2]
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                return False
            
            gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(gray_region)
            # Low confidence threshold to detect any text or pattern
            result = ocr_engine.readtext(enhanced, min_size=5)
            
            if result and len(result) > 0:
                for detection in result:
                    confidence = detection[2]
                    if confidence >= confidence_threshold:
                        print(f"  ✓ Content found: '{detection[1]}' (confidence: {confidence:.3f})")
                        return True
            
            return False
        except Exception as e:
            print(f"  ⚠ OCR error checking content: {e}")
            return False
    
    # First attempt: check for any content
    content_found = check_any_content(content_ocr_region, confidence_threshold=0.1)
    
    if not content_found:
        # Retry: click step36_5_region again and check
        print("  ⚠ No content found, retrying click and check...")
        click_random_in_rect(step36_5_region, "Step 36.5 Region (Retry)")
        wait_time = random.uniform(10, 11)
        print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check again...")
        time.sleep(wait_time)
        
        content_found = check_any_content(content_ocr_region, confidence_threshold=0.1)
        
        if not content_found:
            print("  ❌ No content found after retry, ending task...")
            return False
    
    # Step delay between Step 36 and Step 37
    step_delay(1, 2)
    
    # Helper function for Step 38 "Mark" detection
    def check_mark_text(region_rect, confidence_threshold=0.6):
        """Check if 'Mark' text exists in region with specified confidence threshold"""
        if not initialize_ocr():
            return False
        
        screenshot = take_screenshot()
        if screenshot is None:
            return False
        
        try:
            height, width = screenshot.shape[:2]
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                return False
            
            gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(gray_region)
            result = ocr_engine.readtext(enhanced)
            
            for detection in result:
                text = detection[1].lower()
                confidence = detection[2]
                if 'mark' in text and confidence >= confidence_threshold:
                    print(f"  ✓ 'Mark' text found: '{detection[1]}' (confidence: {confidence:.3f})")
                    return True
            
            return False
        except Exception as e:
            print(f"  ⚠ OCR error checking 'Mark': {e}")
            return False
    
    # Step 37: Click region
    print("\n📍 [Step 37] Clicking region...")
    step37_region = {
        'x_min': min(0.333, 0.697, 0.686, 0.331),  # 0.331
        'x_max': max(0.333, 0.697, 0.686, 0.331),  # 0.697
        'y_min': min(0.552, 0.561, 0.586, 0.581),  # 0.552
        'y_max': max(0.552, 0.561, 0.586, 0.581)   # 0.586
    }
    click_random_in_rect(step37_region, "Step 37 Region")
    time.sleep(1)
    
    # Step delay between Step 37 and Step 38
    step_delay(1, 2)
    
    # Step 38: Click region, wait 5-6 seconds, then OCR for any content
    print("\n📍 [Step 38] Clicking region, waiting, then checking for any content...")
    step38_1_region = {
        'x_min': min(0.633, 0.808, 0.808, 0.622),  # 0.622
        'x_max': max(0.633, 0.808, 0.808, 0.622),  # 0.808
        'y_min': min(0.534, 0.536, 0.557, 0.560),  # 0.534
        'y_max': max(0.534, 0.536, 0.557, 0.560)   # 0.560
    }
    click_random_in_rect(step38_1_region, "Step 38.1 Region")
    
    # Wait 5-6 seconds
    wait_time = random.uniform(5, 6)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check...")
    time.sleep(wait_time)
    
    # OCR region for any content: 左上(354, 1809) 右下(1053, 2116)
    step38_content_ocr_region = pixels_to_rect(354, 1809, 1053, 2116)
    
    # Check for any text or pattern (very low threshold)
    if initialize_ocr():
        screenshot = take_screenshot()
        if screenshot is not None:
            try:
                height, width = screenshot.shape[:2]
                x1 = int(step38_content_ocr_region['x_min'] * width)
                x2 = int(step38_content_ocr_region['x_max'] * width)
                y1 = int(step38_content_ocr_region['y_min'] * height)
                y2 = int(step38_content_ocr_region['y_max'] * height)
                region = screenshot[y1:y2, x1:x2]
                
                if region.size > 0:
                    gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                    enhanced = clahe.apply(gray_region)
                    # Low confidence threshold to detect any text or pattern
                    result = ocr_engine.readtext(enhanced, min_size=5)
                    
                    content_found = False
                    if result and len(result) > 0:
                        for detection in result:
                            confidence = detection[2]
                            if confidence >= 0.1:  # Very low threshold
                                content_found = True
                                print(f"  ✓ Content found: '{detection[1]}' (confidence: {confidence:.3f})")
                                break
                    
                    if content_found:
                        print("  ✓ Content detected, proceeding to next cycle!")
                    else:
                        print("  ⚠ No content found, but continuing to next cycle...")
            except Exception as e:
                print(f"  ⚠ OCR error: {e}, but continuing to next cycle...")
    
    print("\n" + "="*60)
    print("✓ TEST COMPLETE: All steps from Step 33 to Step 38 completed")
    print("="*60 + "\n")
    
    return True


def test_from_step36(row_data):
    """
    Test function: Execute from Step 36 onwards to the end
    This function starts from Step 36 and continues through all remaining steps
    """
    print("\n" + "="*60)
    print("TEST: Starting from Step 36 to End")
    print("="*60 + "\n")
    
    # Step 36: OCR detect any text and execute conditional logic
    print("\n📍 [Step 36] Checking for any text and executing conditional logic...")
    
    # OCR region for text detection: 左上(33, 566) 右下(627, 708)
    text_ocr_region = pixels_to_rect(33, 566, 627, 708)
    
    # Check for any text using OCR (low threshold for detection)
    text_found = False
    if initialize_ocr():
        screenshot = take_screenshot()
        if screenshot is not None:
            try:
                height, width = screenshot.shape[:2]
                x1 = int(text_ocr_region['x_min'] * width)
                x2 = int(text_ocr_region['x_max'] * width)
                y1 = int(text_ocr_region['y_min'] * height)
                y2 = int(text_ocr_region['y_max'] * height)
                region = screenshot[y1:y2, x1:x2]
                
                if region.size > 0:
                    gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                    enhanced = clahe.apply(gray_region)
                    # Low confidence threshold to detect any text
                    result = ocr_engine.readtext(enhanced, min_size=5)
                    
                    if result and len(result) > 0:
                        # If any text detected (low threshold)
                        for detection in result:
                            confidence = detection[2]
                            if confidence >= 0.1:  # Very low threshold
                                text_found = True
                                print(f"  ✓ Text found: '{detection[1]}' (confidence: {confidence:.3f})")
                                break
            except Exception as e:
                print(f"  ⚠ OCR error: {e}")
    
    if text_found:
        # Branch A: If any text found, click region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Branch A] Text found, clicking region...")
        step36_1_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_1_region, "Step 36.1 Region (Text Found)")
        time.sleep(1)
    else:
        # Branch B: If no text found, click three regions in sequence
        print("  🎯 [Branch B] No text found, clicking alternative regions...")
        
        # Step 36.2.1: Click first region 左上(23, 974) 右下(459, 1051)
        print("  🎯 [Step 36.2.1] Clicking first region...")
        step36_2_1_region = pixels_to_rect(23, 974, 459, 1051)
        click_random_in_rect(step36_2_1_region, "Step 36.2.1 Region")
        time.sleep(1)
        
        # Step 36.2.2: Click second region 左上(921, 504) 右下(1034, 559)
        print("  🎯 [Step 36.2.2] Clicking second region...")
        step36_2_2_region = pixels_to_rect(921, 504, 1034, 559)
        click_random_in_rect(step36_2_2_region, "Step 36.2.2 Region")
        time.sleep(1)
        
        # Step 36.2.3: Click third region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Step 36.2.3] Clicking third region...")
        step36_2_3_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_2_3_region, "Step 36.2.3 Region")
        time.sleep(1)
        
        # Step 36.2.4: Click fourth region 左上(197, 2186) 右下(894, 2287)
        print("  🎯 [Step 36.2.4] Clicking fourth region...")
        step36_2_4_region = pixels_to_rect(197, 2186, 894, 2287)
        click_random_in_rect(step36_2_4_region, "Step 36.2.4 Region")
        time.sleep(1)
    
    # Step 36.5: Click final region and check for any text/pattern
    print("  🎯 [Step 36.5] Clicking final region...")
    step36_5_region = {
        'x_min': min(0.344, 0.664, 0.656, 0.350),  # 0.344
        'x_max': max(0.344, 0.664, 0.656, 0.350),  # 0.664
        'y_min': min(0.909, 0.914, 0.949, 0.946),  # 0.909
        'y_max': max(0.909, 0.914, 0.949, 0.946)   # 0.949
    }
    click_random_in_rect(step36_5_region, "Step 36.5 Region (Final)")
    
    # Wait 10-11 seconds
    wait_time = random.uniform(10, 11)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check...")
    time.sleep(wait_time)
    
    # OCR region for any text/pattern: 左上(141, 1802) 右下(1022, 2112)
    content_ocr_region = pixels_to_rect(141, 1802, 1022, 2112)
    
    def check_any_content(region_rect, confidence_threshold=0.1):
        """Check if any text or pattern exists in region (very low threshold)"""
        if not initialize_ocr():
            return False
        
        screenshot = take_screenshot()
        if screenshot is None:
            return False
        
        try:
            height, width = screenshot.shape[:2]
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                return False
            
            gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(gray_region)
            # Low confidence threshold to detect any text or pattern
            result = ocr_engine.readtext(enhanced, min_size=5)
            
            if result and len(result) > 0:
                for detection in result:
                    confidence = detection[2]
                    if confidence >= confidence_threshold:
                        print(f"  ✓ Content found: '{detection[1]}' (confidence: {confidence:.3f})")
                        return True
            
            return False
        except Exception as e:
            print(f"  ⚠ OCR error checking content: {e}")
            return False
    
    # First attempt: check for any content
    content_found = check_any_content(content_ocr_region, confidence_threshold=0.1)
    
    if not content_found:
        # Retry: click step36_5_region again and check
        print("  ⚠ No content found, retrying click and check...")
        click_random_in_rect(step36_5_region, "Step 36.5 Region (Retry)")
        wait_time = random.uniform(10, 11)
        print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check again...")
        time.sleep(wait_time)
        
        content_found = check_any_content(content_ocr_region, confidence_threshold=0.1)
        
        if not content_found:
            print("  ❌ No content found after retry, ending task...")
            return False
    
    # Step delay between Step 36 and Step 37
    step_delay(1, 2)
    
    # Helper function for Step 38 "Mark" detection
    def check_mark_text(region_rect, confidence_threshold=0.6):
        """Check if 'Mark' text exists in region with specified confidence threshold"""
        if not initialize_ocr():
            return False
        
        screenshot = take_screenshot()
        if screenshot is None:
            return False
        
        try:
            height, width = screenshot.shape[:2]
            x1 = int(region_rect['x_min'] * width)
            x2 = int(region_rect['x_max'] * width)
            y1 = int(region_rect['y_min'] * height)
            y2 = int(region_rect['y_max'] * height)
            region = screenshot[y1:y2, x1:x2]
            
            if region.size == 0:
                return False
            
            gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            enhanced = clahe.apply(gray_region)
            result = ocr_engine.readtext(enhanced)
            
            for detection in result:
                text = detection[1].lower()
                confidence = detection[2]
                if 'mark' in text and confidence >= confidence_threshold:
                    print(f"  ✓ 'Mark' text found: '{detection[1]}' (confidence: {confidence:.3f})")
                    return True
            
            return False
        except Exception as e:
            print(f"  ⚠ OCR error checking 'Mark': {e}")
            return False
    
    # Step 37: Click region
    print("\n📍 [Step 37] Clicking region...")
    step37_region = {
        'x_min': min(0.333, 0.697, 0.686, 0.331),  # 0.331
        'x_max': max(0.333, 0.697, 0.686, 0.331),  # 0.697
        'y_min': min(0.552, 0.561, 0.586, 0.581),  # 0.552
        'y_max': max(0.552, 0.561, 0.586, 0.581)   # 0.586
    }
    click_random_in_rect(step37_region, "Step 37 Region")
    time.sleep(1)
    
    # Step delay between Step 37 and Step 38
    step_delay(1, 2)
    
    # Step 38: Click region, wait 5-6 seconds, then OCR for any content
    print("\n📍 [Step 38] Clicking region, waiting, then checking for any content...")
    step38_1_region = {
        'x_min': min(0.633, 0.808, 0.808, 0.622),  # 0.622
        'x_max': max(0.633, 0.808, 0.808, 0.622),  # 0.808
        'y_min': min(0.534, 0.536, 0.557, 0.560),  # 0.534
        'y_max': max(0.534, 0.536, 0.557, 0.560)   # 0.560
    }
    click_random_in_rect(step38_1_region, "Step 38.1 Region")
    
    # Wait 5-6 seconds
    wait_time = random.uniform(5, 6)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds before OCR check...")
    time.sleep(wait_time)
    
    # OCR region for any content: 左上(354, 1809) 右下(1053, 2116)
    step38_content_ocr_region = pixels_to_rect(354, 1809, 1053, 2116)
    
    # Check for any text or pattern (very low threshold)
    if initialize_ocr():
        screenshot = take_screenshot()
        if screenshot is not None:
            try:
                height, width = screenshot.shape[:2]
                x1 = int(step38_content_ocr_region['x_min'] * width)
                x2 = int(step38_content_ocr_region['x_max'] * width)
                y1 = int(step38_content_ocr_region['y_min'] * height)
                y2 = int(step38_content_ocr_region['y_max'] * height)
                region = screenshot[y1:y2, x1:x2]
                
                if region.size > 0:
                    gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                    enhanced = clahe.apply(gray_region)
                    # Low confidence threshold to detect any text or pattern
                    result = ocr_engine.readtext(enhanced, min_size=5)
                    
                    content_found = False
                    if result and len(result) > 0:
                        for detection in result:
                            confidence = detection[2]
                            if confidence >= 0.1:  # Very low threshold
                                content_found = True
                                print(f"  ✓ Content found: '{detection[1]}' (confidence: {confidence:.3f})")
                                break
                    
                    if content_found:
                        print("  ✓ Content detected, proceeding to next cycle!")
                    else:
                        print("  ⚠ No content found, but continuing to next cycle...")
            except Exception as e:
                print(f"  ⚠ OCR error: {e}, but continuing to next cycle...")
    
    print("\n" + "="*60)
    print("✓ TEST COMPLETE: All steps from Step 36 to Step 38 completed")
    print("="*60 + "\n")
    
    return True


def test_from_step20_1(row_data):
    """
    Test function: Execute from Step 20.1 onwards
    This function starts from Step 20.1 (after waiting 5-6 seconds)
    Useful for testing Step 20.1, 20.2, and 20.3
    """
    print("\n" + "="*60)
    print("TEST: Starting from Step 20.1")
    print("="*60 + "\n")
    
    # Step 20.1: Click first region (0.925, 0.068), (0.981, 0.068), (0.975, 0.087), (0.925, 0.089)
    step20_1_region = {
        'x_min': min(0.925, 0.981, 0.975, 0.925),  # 0.925
        'x_max': max(0.925, 0.981, 0.975, 0.925),  # 0.981
        'y_min': min(0.068, 0.068, 0.087, 0.089),  # 0.068
        'y_max': max(0.068, 0.068, 0.087, 0.089)   # 0.089
    }
    print("\n📍 [Step 20.1] Clicking first region...")
    click_random_in_rect(step20_1_region, "Step 20.1 Region")
    time.sleep(1)
    
    # Step 20.2: Click second region (0.531, 0.125), (0.725, 0.125), (0.725, 0.141), (0.536, 0.144)
    step20_2_region = {
        'x_min': min(0.531, 0.725, 0.725, 0.536),  # 0.531
        'x_max': max(0.531, 0.725, 0.725, 0.536),  # 0.725
        'y_min': min(0.125, 0.125, 0.141, 0.144),  # 0.125
        'y_max': max(0.125, 0.125, 0.141, 0.144)   # 0.144
    }
    print("\n📍 [Step 20.2] Clicking second region...")
    click_random_in_rect(step20_2_region, "Step 20.2 Region")
    
    # Step 20.3: Wait 10-12 seconds, then check if region has content (text and images)
    print("\n📍 [Step 20.3] Waiting 10-12 seconds, then checking for content...")
    wait_time = random.uniform(10, 12)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Content check region: (0.033, 0.311) to (0.942, 0.509)
    content_check_region_20_3 = {
        'x_min': min(0.033, 0.961, 0.942, 0.039),  # 0.033
        'x_max': max(0.033, 0.961, 0.942, 0.039),  # 0.961
        'y_min': min(0.311, 0.324, 0.509, 0.510),  # 0.311
        'y_max': max(0.311, 0.324, 0.509, 0.510)   # 0.510
    }
    
    # Check if content is loaded (text and images)
    print(f"  🔍 Checking if region has content (text and images)...")
    content_score, content_loaded = check_region_content_loaded(
        content_check_region_20_3,
        timeout=5,
        check_interval=0.5,
        threshold=0.3  # Lower threshold for faster detection
    )
    
    if not content_loaded:
        # Retry: Repeat Step 20.1 and 20.2
        print(f"  ⚠ Content check failed (score: {content_score:.3f}), retrying Step 20.1 and 20.2...")
        
        # Repeat Step 20.1
        print("  🔄 Repeating Step 20.1...")
        click_random_in_rect(step20_1_region, "Step 20.1 Region (Retry)")
        time.sleep(1)
        
        # Repeat Step 20.2
        print("  🔄 Repeating Step 20.2...")
        click_random_in_rect(step20_2_region, "Step 20.2 Region (Retry)")
        
        # Wait 10-12 seconds again
        wait_time = random.uniform(10, 12)
        print(f"  ⏳ Waiting {wait_time:.1f} seconds before checking again...")
        time.sleep(wait_time)
        
        # Check content again
        print(f"  🔍 Checking content again...")
        content_score, content_loaded = check_region_content_loaded(
            content_check_region_20_3,
            timeout=5,
            check_interval=0.5,
            threshold=0.3
        )
        
        if not content_loaded:
            print(f"  ❌ Content check failed again (score: {content_score:.3f}), canceling task...")
            return False
    
    print(f"  ✓ Content check passed (score: {content_score:.3f}) - proceeding to next step")
    time.sleep(1)
    
    print("\n" + "="*60)
    print("✓ TEST COMPLETE: Step 20.3 passed, ready for Step 21+")
    print("="*60 + "\n")
    
    return True


def test_from_step28_4(row_data):
    """
    Test function: Execute from Step 28.4 onwards to the end
    This function starts from Step 28.4 and continues through all remaining steps
    """
    print("\n" + "="*60)
    print("TEST: Starting from Step 28.4 to End")
    print("="*60 + "\n")
    
    # Step 28.4: Click third region
    # Click region: 左上(47, 801) 右下(354, 912)
    print("  🎯 [Step 28.4] Clicking third region...")
    step28_4_region = pixels_to_rect(47, 801, 354, 912)
    click_random_in_rect(step28_4_region, "Step 28.4 Region")
    time.sleep(1)
    
    # Step delay between Step 28 and Step 29
    step_delay(1, 2)
    
    # Step 29: Click region
    # Click region: 左上(77, 1190) 右下(501, 1248)
    print("\n📍 [Step 29] Clicking region...")
    step29_region = pixels_to_rect(77, 1190, 501, 1248)
    click_random_in_rect(step29_region, "Step 29 Region")
    time.sleep(1)
    
    # Step 29.1: Input Excel I column (Brand)
    print("  ⌨️ [Step 29.1] Inputting Excel I column (Brand)...")
    brand = row_data.get('Brand', '')
    print(f"  ⌨️ Inputting Brand: {brand}")
    input_text_stealth(str(brand), APP_PACKAGE)
    time.sleep(1)
    
    # Step 29.1.1: Click region after Brand input
    # Click region: 左上(96, 2277) 右下(243, 2373)
    print("  🎯 [Step 29.1.1] Clicking region after Brand input...")
    step29_1_1_region = pixels_to_rect(96, 2277, 243, 2373)
    click_random_in_rect(step29_1_1_region, "Step 29.1.1 Region")
    time.sleep(1)
    
    # Step delay between Step 29 and Step 30
    step_delay(1, 2)
    
    # Step 30: Click input region
    print("\n📍 [Step 30] Clicking input region...")
    # Click region: 左上(60, 1411) 右下(617, 1500)
    step30_click_region = pixels_to_rect(60, 1411, 617, 1500)
    click_random_in_rect(step30_click_region, "Step 30 Input Region")
    
    # Wait 1-2 seconds
    wait_time = random.uniform(1, 2)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step delay between Step 30 and Step 33
    step_delay(1, 2)
    
    # Continue with test_from_step33 to execute Step 33 onwards
    return test_from_step33(row_data)


def test_from_step17_1(row_data):
    """
    Test function: Execute from Step 17.1 onwards to the end
    This function starts from Step 17.1 (detecting "Other" text) and continues through all remaining steps
    """
    print("\n" + "="*60)
    print("TEST: Starting from Step 17.1 to End")
    print("="*60 + "\n")
    
    # Step 17.1: Check for "Other" text in specified region (replaced blue button detection)
    def check_other_text(region_rect, timeout=10, check_interval=0.5):
        """Check if "Other" text appears in region
        Args:
            region_rect: Dict with 'x_min', 'x_max', 'y_min', 'y_max'
            timeout: Maximum time to wait (seconds)
            check_interval: Time between checks (seconds)
        Returns:
            True if "Other" text found, False if timeout
        """
        print(f"  🔍 Checking for 'Other' text (timeout: {timeout}s)...")
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            screenshot = take_screenshot()
            if screenshot is None:
                time.sleep(check_interval)
                continue
            
            try:
                height, width = screenshot.shape[:2]
                
                # Extract region
                x1 = int(region_rect['x_min'] * width)
                x2 = int(region_rect['x_max'] * width)
                y1 = int(region_rect['y_min'] * height)
                y2 = int(region_rect['y_max'] * height)
                
                region = screenshot[y1:y2, x1:x2]
                
                if region.size == 0:
                    time.sleep(check_interval)
                    continue
                
                # Use OCR to find "Other" text
                if initialize_ocr():
                    gray_region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                    # Enhance for OCR
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                    enhanced = clahe.apply(gray_region)
                    
                    result = ocr_engine.readtext(enhanced)
                    
                    # Search for "Other" text (case-insensitive)
                    for detection in result:
                        text = detection[1].lower()
                        confidence = detection[2]
                        # Check if text contains "other" (to match "Other", "Othertools", etc.)
                        if 'other' in text and confidence > 0.5:
                            print(f"  ✓ 'Other' text detected: '{detection[1]}' (confidence: {confidence:.3f})")
                            return True
                
            except Exception as e:
                print(f"  ⚠ 'Other' text check error: {e}")
            
            time.sleep(check_interval)
        
        print(f"  ❌ 'Other' text not found within {timeout}s")
        return False
    
    # "Other" text detection region: (0.031, 0.450) to (0.036, 0.522) to (0.950, 0.517) to (0.939, 0.451)
    other_text_detection_region = {
        'x_min': min(0.031, 0.036, 0.950, 0.939),  # 0.031
        'x_max': max(0.031, 0.036, 0.950, 0.939),  # 0.950
        'y_min': min(0.450, 0.522, 0.517, 0.451),  # 0.450
        'y_max': max(0.450, 0.522, 0.517, 0.451)   # 0.522
    }
    
    # Blue button click region: (0.358, 0.279) to (0.367, 0.316) to (0.644, 0.281) to (0.656, 0.314)
    blue_button_click_region = {
        'x_min': min(0.358, 0.367, 0.644, 0.656),  # 0.358
        'x_max': max(0.358, 0.367, 0.644, 0.656),  # 0.656
        'y_min': min(0.279, 0.316, 0.281, 0.314),  # 0.279
        'y_max': max(0.279, 0.316, 0.281, 0.314)   # 0.316
    }
    
    # Final submit region (needed for retry logic)
    final_submit_region = {
        'x_min': min(0.125, 0.894, 0.903, 0.150),  # 0.125
        'x_max': max(0.125, 0.894, 0.903, 0.150),  # 0.903
        'y_min': min(0.906, 0.915, 0.956, 0.951),  # 0.906
        'y_max': max(0.906, 0.915, 0.956, 0.951)   # 0.956
    }
    
    # First attempt: check for "Other" text
    print("\n📍 [Step 17.1] Checking for 'Other' text...")
    other_text_found = check_other_text(other_text_detection_region, timeout=10, check_interval=0.5)
    
    if other_text_found:
        print("  ✓ 'Other' text found, clicking blue button region...")
        click_random_in_rect(blue_button_click_region, "Blue Button")
    else:
        # Retry Step 17 if "Other" text not found
        print("  ⚠ 'Other' text not found, retrying Step 17...")
        # Click final submit button again
        click_random_in_rect(final_submit_region, "Final Submit Button (Retry)")
        
        # Wait 13-15 seconds again
        wait_time = random.uniform(13, 15)
        print(f"  ⏳ Waiting {wait_time:.1f} seconds before checking again...")
        time.sleep(wait_time)
        
        # Check for "Other" text again
        other_text_found = check_other_text(other_text_detection_region, timeout=10, check_interval=0.5)
        
        if other_text_found:
            print("  ✓ 'Other' text found on retry, clicking blue button region...")
            click_random_in_rect(blue_button_click_region, "Blue Button (Retry)")
        else:
            print("  ⚠ 'Other' text still not found after retry, continuing anyway...")
    
    # Continue with remaining steps from Step 18 onwards
    # (Copy all steps from Step 18 to Step 38 from upload_images_phase function)
    # For brevity, I'll call the existing test_from_step21 function but start from Step 18
    # Actually, let me extract the steps properly...
    
    # Step delay between Step 17 and Step 18
    step_delay(1, 2)
    
    # Step 18: Click region (0.444, 0.274), (0.544, 0.273), (0.547, 0.286), (0.464, 0.284)
    print("\n📍 [Step 18] Clicking region...")
    step18_region = {
        'x_min': min(0.444, 0.544, 0.547, 0.464),  # 0.444
        'x_max': max(0.444, 0.544, 0.547, 0.464),  # 0.547
        'y_min': min(0.274, 0.273, 0.286, 0.284),  # 0.273
        'y_max': max(0.274, 0.273, 0.286, 0.284)   # 0.286
    }
    click_random_in_rect(step18_region, "Step 18 Region")
    time.sleep(1)
    
    # Step delay between Step 18 and Step 19
    step_delay(1, 2)
    
    # Step 19: Click region (0.053, 0.329), (0.456, 0.324), (0.453, 0.505), (0.078, 0.504)
    print("\n📍 [Step 19] Clicking region...")
    step19_region = {
        'x_min': min(0.053, 0.456, 0.453, 0.078),  # 0.053
        'x_max': max(0.053, 0.456, 0.453, 0.078),  # 0.456
        'y_min': min(0.329, 0.324, 0.505, 0.504),  # 0.324
        'y_max': max(0.329, 0.324, 0.505, 0.504)   # 0.505
    }
    click_random_in_rect(step19_region, "Step 19 Region")
    time.sleep(1)
    
    # Step delay between Step 19 and Step 20
    step_delay(1, 2)
    
    # Step 20: Wait 5-6 seconds, then click two regions
    print("\n📍 [Step 20] Waiting 5-6 seconds, then clicking regions...")
    wait_time = random.uniform(5, 6)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Step 20.1: Click first region (0.925, 0.068), (0.981, 0.068), (0.975, 0.087), (0.925, 0.089)
    step20_1_region = {
        'x_min': min(0.925, 0.981, 0.975, 0.925),  # 0.925
        'x_max': max(0.925, 0.981, 0.975, 0.925),  # 0.981
        'y_min': min(0.068, 0.068, 0.087, 0.089),  # 0.068
        'y_max': max(0.068, 0.068, 0.087, 0.089)   # 0.089
    }
    print("  🎯 Clicking first region...")
    click_random_in_rect(step20_1_region, "Step 20.1 Region")
    time.sleep(1)
    
    # Step 20.2: Click second region (0.531, 0.125), (0.725, 0.125), (0.725, 0.141), (0.536, 0.144)
    step20_2_region = {
        'x_min': min(0.531, 0.725, 0.725, 0.536),  # 0.531
        'x_max': max(0.531, 0.725, 0.725, 0.536),  # 0.725
        'y_min': min(0.125, 0.125, 0.141, 0.144),  # 0.125
        'y_max': max(0.125, 0.125, 0.141, 0.144)   # 0.144
    }
    print("  🎯 Clicking second region...")
    click_random_in_rect(step20_2_region, "Step 20.2 Region")
    
    # Step 20.3: Wait 10-12 seconds, then check if region has content (text and images)
    print("\n📍 [Step 20.3] Waiting 10-12 seconds, then checking for content...")
    wait_time = random.uniform(10, 12)
    print(f"  ⏳ Waiting {wait_time:.1f} seconds...")
    time.sleep(wait_time)
    
    # Content check region: (0.033, 0.311) to (0.942, 0.509)
    content_check_region_20_3 = {
        'x_min': min(0.033, 0.961, 0.942, 0.039),  # 0.033
        'x_max': max(0.033, 0.961, 0.942, 0.039),  # 0.961
        'y_min': min(0.311, 0.324, 0.509, 0.510),  # 0.311
        'y_max': max(0.311, 0.324, 0.509, 0.510)   # 0.510
    }
    
    # Check if content is loaded (text and images)
    print(f"  🔍 Checking if region has content (text and images)...")
    content_score, content_loaded = check_region_content_loaded(
        content_check_region_20_3,
        timeout=5,
        check_interval=0.5,
        threshold=0.3  # Lower threshold for faster detection
    )
    
    if not content_loaded:
        # Retry: Repeat Step 20.1 and 20.2
        print(f"  ⚠ Content check failed (score: {content_score:.3f}), retrying Step 20.1 and 20.2...")
        
        # Repeat Step 20.1
        print("  🔄 Repeating Step 20.1...")
        click_random_in_rect(step20_1_region, "Step 20.1 Region (Retry)")
        time.sleep(1)
        
        # Repeat Step 20.2
        print("  🔄 Repeating Step 20.2...")
        click_random_in_rect(step20_2_region, "Step 20.2 Region (Retry)")
        
        # Wait 10-12 seconds again
        wait_time = random.uniform(10, 12)
        print(f"  ⏳ Waiting {wait_time:.1f} seconds before checking again...")
        time.sleep(wait_time)
        
        # Check content again
        print(f"  🔍 Checking content again...")
        content_score, content_loaded = check_region_content_loaded(
            content_check_region_20_3,
            timeout=5,
            check_interval=0.5,
            threshold=0.3
        )
        
        if not content_loaded:
            print(f"  ❌ Content check failed again (score: {content_score:.3f}), canceling task...")
            return False
    
    print(f"  ✓ Content check passed (score: {content_score:.3f}) - proceeding to next step")
    time.sleep(1)
    
    # Now call test_from_step21 to continue from Step 21 onwards
    # But we need to extract the code from Step 21 to Step 38
    # For now, let's call the existing function but we need to modify it
    # Actually, let's just continue with the steps inline
    
    # Step delay between Step 20 and Step 21
    step_delay(1, 2)
    
    # Continue with test_from_step21 logic (Step 21 onwards)
    # We'll call it but it expects to start from Step 21, which is what we want
    return test_from_step21(row_data)


if __name__ == "__main__":
    import sys
    
    # Check if test mode is requested
    if len(sys.argv) > 1 and sys.argv[1] == "--test-step21":
        # Test mode: Execute from Step 21
        print("\n" + "="*60)
        print("TEST MODE: Starting from Step 21 to End")
        print("="*60 + "\n")
        
        # Create a dummy row_data for testing
        test_row_data = {
            'SKU': 'TEST',
            'ProductNameCn': '测试产品',
            'ProductNameEn': 'Test Product',
            'GenderEn': 'men',
            'HKPrice': '100',
            'MYPrice': '50',
            'SGPrice': '80',
            'Brand': 'Test Brand',
            'ImageFolder': '',
            'Description': 'Test Description'
        }
        
        try:
            test_from_step21(test_row_data)
        except KeyboardInterrupt:
            print("\n\nTest interrupted by user")
        except Exception as e:
            print(f"\n\nTest error: {e}")
            import traceback
            traceback.print_exc()
    elif len(sys.argv) > 1 and sys.argv[1] == "--test-step20-1":
        # Test mode: Execute from Step 20.1
        print("\n" + "="*60)
        print("TEST MODE: Starting from Step 20.1")
        print("="*60 + "\n")
        
        # Create a dummy row_data for testing
        test_row_data = {
            'SKU': 'TEST',
            'ProductNameCn': '测试产品',
            'ProductNameEn': 'Test Product',
            'GenderEn': 'men',
            'HKPrice': '100',
            'MYPrice': '50',
            'SGPrice': '80',
            'Brand': 'Test Brand',
            'ImageFolder': '',
            'Description': 'Test Description'
        }
        
        try:
            test_from_step20_1(test_row_data)
        except KeyboardInterrupt:
            print("\n\nTest interrupted by user")
        except Exception as e:
            print(f"\n\nTest error: {e}")
            import traceback
            traceback.print_exc()
    elif len(sys.argv) > 1 and sys.argv[1] == "--test-step33":
        # Test mode: Execute from Step 33
        print("\n" + "="*60)
        print("TEST MODE: Starting from Step 33 to End")
        print("="*60 + "\n")
        
        # Create a dummy row_data for testing
        test_row_data = {
            'SKU': 'TEST',
            'ProductNameCn': '测试产品',
            'ProductNameEn': 'Test Product',
            'GenderEn': 'men',
            'HKPrice': '100',
            'MYPrice': '50',
            'SGPrice': '80',
            'Brand': 'Test Brand',
            'ImageFolder': '',
            'Description': 'Test Description'
        }
        
        try:
            test_from_step33(test_row_data)
        except KeyboardInterrupt:
            print("\n\nTest interrupted by user")
        except Exception as e:
            print(f"\n\nTest error: {e}")
            import traceback
            traceback.print_exc()
    elif len(sys.argv) > 1 and sys.argv[1] == "--test-step17-1":
        # Test mode: Execute from Step 17.1
        print("\n" + "="*60)
        print("TEST MODE: Starting from Step 17.1 to End")
        print("="*60 + "\n")
        
        # Create a dummy row_data for testing
        test_row_data = {
            'SKU': 'TEST',
            'ProductNameCn': '测试产品',
            'ProductNameEn': 'Test Product',
            'GenderEn': 'men',
            'HKPrice': '100',
            'MYPrice': '50',
            'SGPrice': '80',
            'Brand': 'Test Brand',
            'ImageFolder': '',
            'Description': 'Test Description'
        }
        
        try:
            test_from_step17_1(test_row_data)
        except KeyboardInterrupt:
            print("\n\nTest interrupted by user")
        except Exception as e:
            print(f"\n\nTest error: {e}")
            import traceback
            traceback.print_exc()
    elif len(sys.argv) > 1 and sys.argv[1] == "--test-step28-4":
        # Test mode: Execute from Step 28.4
        print("\n" + "="*60)
        print("TEST MODE: Starting from Step 28.4 to End")
        print("="*60 + "\n")
        
        # Create a dummy row_data for testing
        test_row_data = {
            'SKU': 'TEST',
            'ProductNameCn': '测试产品',
            'ProductNameEn': 'Test Product',
            'GenderEn': 'men',
            'HKPrice': '100',
            'MYPrice': '50',
            'SGPrice': '80',
            'Brand': 'Test Brand',
            'ImageFolder': '',
            'Description': 'Test Description'
        }
        
        try:
            test_from_step28_4(test_row_data)
        except KeyboardInterrupt:
            print("\n\nTest interrupted by user")
        except Exception as e:
            print(f"\n\nTest error: {e}")
            import traceback
            traceback.print_exc()
    elif len(sys.argv) > 1 and sys.argv[1] == "--test-step36":
        # Test mode: Execute from Step 36
        print("\n" + "="*60)
        print("TEST MODE: Starting from Step 36 to End")
        print("="*60 + "\n")
        
        # Create a dummy row_data for testing
        test_row_data = {
            'SKU': 'TEST',
            'ProductNameCn': '测试产品',
            'ProductNameEn': 'Test Product',
            'GenderEn': 'men',
            'HKPrice': '100',
            'MYPrice': '50',
            'SGPrice': '80',
            'Brand': 'Test Brand',
            'ImageFolder': '',
            'Description': 'Test Description'
        }
        
        try:
            test_from_step36(test_row_data)
        except KeyboardInterrupt:
            print("\n\nTest interrupted by user")
        except Exception as e:
            print(f"\n\nTest error: {e}")
            import traceback
            traceback.print_exc()
    else:
        # Normal mode: Execute full main function
        try:
            main()
        except KeyboardInterrupt:
            print("\n\nScript interrupted by user")
        except Exception as e:
            print(f"\n\nFatal error: {e}")
            import traceback
            traceback.print_exc()
